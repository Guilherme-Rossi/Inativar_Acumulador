"""
Utilitários para trabalhar com planilhas Excel.

Carrega empresas, acumuladores, e aplica regras de negócio.
Retorna objetos tipados da camada de domínio.
"""

import re
import time
from pathlib import Path
from typing import List, Optional, Set, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    import win32com.client as win32
except Exception:
    win32 = None

from agente_rpa.config.settings import CFG
from agente_rpa.core.logger import log
from agente_rpa.dominio.modelos import Empresa, Acumulador


def normalizar_codigo(valor) -> Optional[int]:
    """Converte valor para inteiro, tratando NaN e strings."""
    if pd.isna(valor):
        return None
    try:
        texto = str(valor).strip()
        if not texto:
            return None
        return int(float(texto.replace(".0", "")))
    except Exception:
        return None


def carregar_empresas(caminho_planilha: Path) -> List[Empresa]:
    df = pd.read_excel(caminho_planilha)
    colunas = {str(c).lower().strip(): c for c in df.columns}

    coluna_codigo = next(
        (colunas[n] for n in ("codigo", "código", "cod", "coluna a") if n in colunas),
        df.columns[0]
    )
    coluna_empresa = next(
        (colunas[n] for n in ("empresa", "nome", "razao social", "razão social") if n in colunas),
        None
    )

    empresas: List[Empresa] = []
    for _, row in df.iterrows():
        codigo = normalizar_codigo(row[coluna_codigo])
        if codigo is None:
            continue
        nome = str(row[coluna_empresa]).strip() if coluna_empresa else ""
        empresas.append(Empresa(codigo=codigo, nome=nome))

    return empresas


def carregar_empresas_ja_inativadas_por_cor(caminho_planilha: Path) -> Set[int]:
    """Retorna códigos de empresa marcados em amarelo na planilha de controle."""
    if not caminho_planilha.exists():
        return set()

    amarelos = {
        "FFFFFF00",
        "00FFFF00",
        "FFFF00",
        "FFFFFF99",
        "FFFFEB9C",
        "00FFEB9C",
    }

    try:
        caminho = converter_xls_para_xlsx_se_preciso(caminho_planilha)
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active

        empresas_amarelas: Set[int] = set()
        for row in ws.iter_rows(min_row=2):
            codigo = normalizar_codigo(row[0].value)
            if codigo is None:
                continue

            linha_marcada = False
            for cell in row:
                fill = cell.fill
                if not fill or fill.fill_type != "solid":
                    continue

                cor = ""
                if getattr(fill.start_color, "rgb", None):
                    cor = fill.start_color.rgb.upper()
                elif getattr(fill.start_color, "index", None):
                    cor = str(fill.start_color.index).upper()

                if cor in amarelos:
                    linha_marcada = True
                    break

            if linha_marcada:
                empresas_amarelas.add(codigo)

        return empresas_amarelas
    except Exception as exc:
        log.info(f"[AVISO] Falha ao ler planilha de empresas já inativadas: {exc}")
        return set()


def carregar_acumuladores_principais(
    caminho_planilha: Path,
    faixa_min: int = 1,
    faixa_max: int = 211,
    excluidos: tuple[int, ...] = (2,),
    faixa_analise_min: int = 212,
    faixa_analise_max: int = 1000,
) -> List[Acumulador]:
    wb = load_workbook(caminho_planilha, data_only=True)
    ws = wb.active

    if ws.max_row < 2 or ws.max_column < 1:
        raise ValueError("Planilha de acumuladores está vazia ou sem dados.")

    amarelos = {"FFFFFF00", "00FFFF00", "FFFF00", "FFFFFF99", "FFFFEB9C", "00FFEB9C"}
    acumuladores: List[Acumulador] = []

    for row in ws.iter_rows(min_row=2):
        codigo = normalizar_codigo(row[0].value)
        if codigo is None:
            continue

        fill = row[0].fill
        rgb = None
        if fill and fill.fill_type == "solid":
            rgb = getattr(fill.start_color, "rgb", None)
            if rgb:
                rgb = rgb.upper()

        marcado_amarelo = rgb in amarelos
        deve_inativar = (
            faixa_min <= codigo <= faixa_max
            and codigo not in excluidos
            and marcado_amarelo
        )
        deve_marcar_analise = faixa_analise_min <= codigo <= faixa_analise_max

        acumuladores.append(Acumulador(
            codigo=codigo,
            deve_inativar=deve_inativar,
            deve_marcar_analise=deve_marcar_analise,
        ))

    return acumuladores


def verificar_relatorio_existente(codigo_empresa: int, sufixo: str = "") -> Optional[Path]:
    nome_base = f"RELAÇÃO DE ACUMULADORES - {codigo_empresa}{sufixo}"
    caminho_xlsx = CFG.pasta_relatorios / f"{nome_base}.xlsx"
    caminho_xls = CFG.pasta_relatorios / f"{nome_base}.xls"

    if caminho_xlsx.exists():
        return caminho_xlsx
    if caminho_xls.exists():
        return caminho_xls
    return None


def arquivo_xls_valido(caminho: Path) -> bool:
    try:
        with open(caminho, "rb") as arquivo:
            cabecalho = arquivo.read(8)
        if caminho.suffix.lower() == ".xlsx":
            return cabecalho[:4] == b"PK\x03\x04"
        return cabecalho[:4] in (b"\xd0\xcf\x11\xe0", b"\x09\x08\x10\x00", b"\x01\x02\x06\x00")
    except Exception:
        return False


def converter_xls_para_xlsx_se_preciso(caminho_arquivo: Path) -> Path:
    if caminho_arquivo.suffix.lower() != ".xls":
        return caminho_arquivo

    destino = caminho_arquivo.with_suffix(".xlsx")
    if destino.exists() and arquivo_xls_valido(destino):
        return destino

    if win32 is None:
        raise RuntimeError("win32com não está disponível para converter .xls para .xlsx.")

    log.info("[INFO] Convertendo arquivo .xls para .xlsx via Excel.")
    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(caminho_arquivo.resolve()))
        wb.SaveAs(str(destino.resolve()), FileFormat=51)
        wb.Close()
        excel.Quit()
        return destino
    except Exception as exc:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        raise RuntimeError(f"Falha ao converter arquivo usando Excel: {exc}")


def aguardar_relatorio_empresa(codigo_empresa: int, timeout: int = 15, sufixo: str = "") -> Path:
    for _ in range(timeout):
        caminho = verificar_relatorio_existente(codigo_empresa, sufixo=sufixo)
        if caminho is not None:
            return caminho
        time.sleep(1)
    raise FileNotFoundError(f"O Excel 'RELAÇÃO DE ACUMULADORES - {codigo_empresa}{sufixo}' não apareceu na pasta.")


def processar_relatorio_extraido(caminho_arquivo: Path) -> List[int]:
    caminho = converter_xls_para_xlsx_se_preciso(caminho_arquivo)
    df = pd.read_excel(caminho, header=None)
    df = df.iloc[6:].reset_index(drop=True)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    acumuladores: List[int] = []
    for valor in df.iloc[:, 0].tolist():
        codigo = normalizar_codigo(valor)
        if codigo is not None:
            acumuladores.append(codigo)
    return acumuladores


def colorir_planilha_verificacao(caminho_planilha: Path, codigos_inativados: Set[int]) -> None:
    caminho = converter_xls_para_xlsx_se_preciso(caminho_planilha)
    wb = load_workbook(caminho)
    ws = wb.active

    cor_verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    cor_amarela = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=1):
        celula_codigo = row[0]
        codigo = normalizar_codigo(celula_codigo.value)
        if codigo is None:
            continue

        if codigo in codigos_inativados:
            for cell in row:
                cell.fill = cor_verde
        elif CFG.faixa_marcacao_min <= codigo <= CFG.faixa_marcacao_max:
            for cell in row:
                cell.fill = cor_amarela

    wb.save(caminho)
    log.info("[INFO] Planilha de verificação colorida: verde=inativados, amarelo=possíveis/análise.")
