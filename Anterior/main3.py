from __future__ import annotations

import csv
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32

import pytesseract
from pytesseract import Output

try:
    import pyautogui
except Exception:
    pyautogui = None

from agente_dominio import AgenteConfig, AgenteDominio


# ============================================================
# EXCEÇÕES PERSONALIZADAS DO PROJETO
# ============================================================
class SemAcumuladoresException(Exception):
    """ Erro disparado de propósito quando a empresa não possui nenhum acumulador cadastrado """
    pass


# ============================================================
# CONFIGURACAO UNIFICADA
# ============================================================
@dataclass(frozen=True)
class Config:
    base_dir: Path = Path.home() / "Desktop" / "PYTHON DOCS" / "Projects Agrelli" / "P1 - Inativar Acumulador"
    planilha_empresas: Path = base_dir / "EMPRESA PARA INATIVAR ACUMULADOR.xlsx"
    planilha_acumuladores: Path = base_dir / "RELAÇÃO DE ACUMULADORES.xlsx"
    pasta_relatorios: Path = base_dir / "Relação Por empresa"
    pasta_logs: Path = base_dir / "Relatorio Final"
    pasta_screenshots: Path = pasta_logs / "screenshots"
    arquivo_execucao_csv: Path = pasta_logs / "execucao_consolidada.csv"
    pasta_imagens: Path = base_dir / "imagens_rpa"
    pasta_debug_agente: Path = pasta_logs / "debug_agente"
    
    pasta_inativacao: Path = base_dir / "Relatório Inativação"

    faixa_min: int = 212
    faixa_max: int = 1000
   
   
    dry_run: bool = False  
    
    data_inativacao: str = "32026"

    caminho_tesseract: str = r'C:\Users\guilherme.rossi\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
    
    regiao_acumuladores: tuple = (545, 250, 829, 614)
    
    pyautogui_pause: float = 0.5
    tempo_excel: float = 4.0
    tempo_dominio_baixo: float = 7.5
    tempo_dominio: float = 16.0
    tempo_dominio_max: float = 27.0 

    retry_tentativas: int = 3
    retry_espera: float = 2.0

    esc_loop_qtd: int = 6
    esc_loop_intervalo: float = 0.3
    segundos_iniciais: int = 5

    x_cancelar: int = 1316 
    y_cancelar: int = 304
    
    x_foco_listagem: int = 1527
    y_foco_listagem: int = 308
    
    x_excel: int = 29
    y_excel: int = 302
    x_nome_arquivo: int = 359
    y_nome_arquivo: int = 418
    
    x_salvar: int = 603
    y_salvar: int = 414
    
    x_voltar_dominio: int = 559
    y_voltar_dominio: int = 581
    
    x_buscar_campo: int = 1694
    y_buscar_campo: int = 700
    x_buscar_botao: int = 1700
    y_buscar_botao: int = 699
    x_segundo_codigo_1: int = 1525
    y_segundo_codigo_1: int = 351
    x_situacao: int = 932
    y_situacao: int = 349
    
    # NOVAS COORDENADAS PARA TRATAMENTO DE IMPOSTOS
    x_aba_impostos: int = 741
    y_aba_impostos: int = 462
    x_excluir_imposto: int = 1176 # Centro (1136+1216)/2
    y_excluir_imposto: int = 770  # Centro (761+781)/2
    
    x_inativo: int = 922
    y_inativo: int = 378
    x_data: int = 1192
    y_data: int = 356
    x_gravar: int = 1326
    y_gravar: int = 370


CFG = Config()

def validar_ambiente() -> None:
    if pyautogui is None:
        raise RuntimeError("Biblioteca 'pyautogui' não encontrada.")
    
    caminho_tess = Path(CFG.caminho_tesseract)
    if not caminho_tess.exists():
        raise FileNotFoundError(f"EXECUTÁVEL TESSERACT NÃO ENCONTRADO em: {CFG.caminho_tesseract}")
    
    pytesseract.pytesseract.tesseract_cmd = str(caminho_tess)

    CFG.pasta_relatorios.mkdir(parents=True, exist_ok=True)
    CFG.pasta_logs.mkdir(parents=True, exist_ok=True)
    CFG.pasta_screenshots.mkdir(parents=True, exist_ok=True)
    CFG.pasta_imagens.mkdir(parents=True, exist_ok=True)
    CFG.pasta_debug_agente.mkdir(parents=True, exist_ok=True)
    CFG.pasta_inativacao.mkdir(parents=True, exist_ok=True) 


class Logger:
    def __init__(self) -> None:
        self.caminho = CFG.pasta_logs / "execucao_geral.log"

    def info(self, mensagem: str) -> None:
        linha = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {mensagem}"
        print(linha)
        with open(self.caminho, "a", encoding="utf-8") as f:
            f.write(linha + "\n")


LOG = Logger()


def executar_com_retry(func: Callable, tentativas: int = None, espera: float = None, descricao: str = "Operacao"):
    if tentativas is None: tentativas = CFG.retry_tentativas
    if espera is None: espera = CFG.retry_espera

    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            return func()
        except SemAcumuladoresException as exc_vazio:
            raise exc_vazio
        except Exception as exc:
            ultimo_erro = exc
            LOG.info(f"[RETRY] {descricao} falhou na tentativa {tentativa}/{tentativas}: {exc}")
            if tentativa < tentativas:
                time.sleep(espera)
    raise ultimo_erro


def normalizar_codigo(valor) -> Optional[int]:
    if pd.isna(valor): return None
    try:
        texto = str(valor).strip()
        if not texto: return None
        texto = texto.replace(".0", "")
        return int(float(texto))
    except Exception:
        return None


def criar_agente() -> AgenteDominio:
    return AgenteDominio(
        AgenteConfig(
            pasta_imagens=CFG.pasta_imagens,
            confidence=0.80,
            grayscale=False, 
            pyautogui_pause=CFG.pyautogui_pause,
            tempo_padrao=CFG.tempo_dominio,
            tempo_max=CFG.tempo_dominio_max,
            usar_imagem=True,
            salvar_screenshots_debug=True,
            pasta_debug=CFG.pasta_debug_agente,
        )
    )

def carregar_empresas(caminho: Path) -> pd.DataFrame:
    df = pd.read_excel(caminho)
    colunas = {c.lower().strip(): c for c in df.columns}

    coluna_codigo = next((colunas[n] for n in ("codigo", "código", "cod", "coluna a") if n in colunas), df.columns[0])
    coluna_empresa = next((colunas[n] for n in ("empresa", "nome", "razao social", "razão social") if n in colunas), None)

    df = df.copy()
    df["CODIGO_EMPRESA"] = df[coluna_codigo].apply(normalizar_codigo)
    df["NOME_EMPRESA"] = df[coluna_empresa].astype(str).str.strip() if coluna_empresa else ""

    df = df[df["CODIGO_EMPRESA"].notna()].copy()
    df["CODIGO_EMPRESA"] = df["CODIGO_EMPRESA"].astype(int)
    
    return df[["CODIGO_EMPRESA", "NOME_EMPRESA"]].drop_duplicates().sort_values("CODIGO_EMPRESA")

def cor_para_rgb(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid": return None
    if getattr(fill.start_color, "rgb", None): return fill.start_color.rgb.upper()
    if getattr(fill.start_color, "index", None): return str(fill.start_color.index).upper()
    return None

def carregar_acumuladores_principais(caminho_planilha: Path) -> list[dict]:
    wb = load_workbook(caminho_planilha, data_only=True)
    ws = wb.active
    
    if ws.max_row < 2 or ws.max_column < 1:
        raise ValueError("Planilha de acumuladores está vazia ou sem dados.")

    amarelos = {"FFFFFF00", "00FFFF00", "FFFF00", "FFFFFF99", "FFFFEB9C", "00FFEB9C"}

    registros =[]
    for row in ws.iter_rows(min_row=2):
        codigo = normalizar_codigo(row[0].value)
        if codigo is None: continue
        
        rgb = cor_para_rgb(row[0])
        marcado_amarelo = "SIM" if rgb in amarelos else "NAO"
        
        registros.append({
            "ACUMULADOR": codigo,
            "AMARELO_PLANILHA": marcado_amarelo,
            "DENTRO_FAIXA": "SIM" if CFG.faixa_min <= CFG.faixa_max else "NAO"
        })
    return registros

def verificar_relatorio_existente(codigo_empresa: int) -> Optional[Path]:
    nome_base = f"RELAÇÃO DE ACUMULADORES - {codigo_empresa}"
    caminho_xls = CFG.pasta_relatorios / f"{nome_base}.xls"
    caminho_xlsx = CFG.pasta_relatorios / f"{nome_base}.xlsx"

    if caminho_xlsx.exists():
        return caminho_xlsx
    if caminho_xls.exists():
        return caminho_xls
    return None

def aguardar_relatorio_empresa(codigo_empresa: int, timeout: int = 15) -> Path:
    caminho_esperado = CFG.pasta_relatorios / f"RELAÇÃO DE ACUMULADORES - {codigo_empresa}.xls"
    for _ in range(timeout):
        if caminho_esperado.exists():
            return caminho_esperado
        time.sleep(1)
    raise FileNotFoundError(f"O Excel '{caminho_esperado.name}' não apareceu na pasta.")

def arquivo_xls_valido(caminho: Path) -> bool:
    try:
        with open(caminho, "rb") as f:
            cabecalho = f.read(8)
        if caminho.suffix == ".xlsx":
            return cabecalho[:4] == b'PK\x03\x04'
        return cabecalho[:4] in (b'\xd0\xcf\x11\xe0', b'\x09\x08\x10\x00', b'\x01\x02\x06\x00')
    except Exception:
        return False

def processar_relatorio_extraido(caminho_arquivo: Path) -> list[int]:
    if caminho_arquivo.suffix == ".xls":
        destino = caminho_arquivo.with_suffix(".xlsx")
        
        if not destino.exists():
            LOG.info(f"[INFO] Convertendo formato nativo do Domínio para .xlsx moderno via Excel...")
            try:
                excel = win32.DispatchEx('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(str(caminho_arquivo.resolve()))
                wb.SaveAs(str(destino.resolve()), FileFormat=51)
                wb.Close()
                excel.Quit()
            except Exception as e:
                try: excel.Quit() 
                except: pass
                raise RuntimeError(f"Falha ao tentar converter o arquivo usando o Excel: {e}")
                
        caminho_arquivo = destino

    df = pd.read_excel(caminho_arquivo, header=None)
    df = df.iloc[6:].reset_index(drop=True)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    acumuladores =[]
    if not df.empty:
        for valor in df.iloc[:, 0].tolist():
            codigo = normalizar_codigo(valor)
            if codigo is not None:
                acumuladores.append(codigo) 
    return acumuladores

def clicar_texto_ocr(texto_alvo: str) -> bool:
    LOG.info(f"[OCR] Procurando a palavra '{texto_alvo}' para clicar...")
    try:
        imagem_tela = pyautogui.screenshot(region=CFG.regiao_acumuladores)
        dados_ocr = pytesseract.image_to_data(imagem_tela, lang='por', output_type=Output.DICT)
        
        for i in range(len(dados_ocr['text'])):
            texto_lido = dados_ocr['text'][i].strip()
            conf_raw = str(dados_ocr["conf"][i]).strip()

            try:
                confianca = float(conf_raw)
            except Exception:
                confianca = -1
            
            if texto_alvo.lower() in texto_lido.lower() and confianca > 50:
                x_crop = dados_ocr['left'][i] + (dados_ocr['width'][i] // 2)
                y_crop = dados_ocr['top'][i] + (dados_ocr['height'][i] // 2)
                
                x_tela = CFG.regiao_acumuladores[0] + x_crop
                y_tela = CFG.regiao_acumuladores[1] + y_crop
                
                LOG.info(f"[OCR] Palavra '{texto_alvo}' encontrada! Clicando em X:{x_tela} Y:{y_tela}")
                pyautogui.click(x_tela, y_tela)
                return True
    except Exception:
        pass
    return False


# ============================================================
# RPA FLOW (MÓDULOS DE EXTRAÇÃO E INATIVAÇÃO)
# ============================================================

def entrar_em_acumuladores_por_atalho(agente: AgenteDominio) -> None:
    max_tentativas = 3
    tela_aberta = False
    cadastro_vazio = False

    for tentativa in range(max_tentativas):
        LOG.info(f"[INFO] Tentando entrar em Acumuladores (Alt+A+A) - Tentativa {tentativa+1}/{max_tentativas}")
        
        if tentativa > 0:
            if agente.localizar_imagem("tela_inteira.png", confidence=0.7) or \
               agente.localizar_imagem("tela_acumuladores.png", confidence=0.7) or \
               agente.localizar_imagem("btn_cancelar2.png", confidence=0.7) or \
               agente.localizar_imagem("tela_cancelar2.png", confidence=0.7) or \
               agente.localizar_imagem("tela_novo.png", confidence=0.7) or \
               agente.localizar_imagem("btn_novo2.png", confidence=0.7):
                LOG.info("[VISÃO] A tela abriu com atraso! Não vamos dar ESC.")
                tela_aberta = True
                break
                
            LOG.info("[INFO] Limpando a tela com ESCs antes de tentar de novo...")
            agente.fechar_sistema_para_loop(quantidade_esc=3, intervalo=0.2)
            
        if not tela_aberta:
            agente.sequencia_alt_aa(seg_entre_teclas=0.5, seg_final=1.0)
            agente.pressionar("enter", 0.5)

        LOG.info(f"[INFO] Aguardando a tela do Domínio abrir (até {CFG.tempo_dominio_max}s)...")
        tempo_espera = time.time()
        
        while (time.time() - tempo_espera) < CFG.tempo_dominio_max:
            if agente.localizar_imagem("tela_aviso1.png", confidence=0.7):
                LOG.info("[VISÃO] Aviso de 'cadastro vazio' detectado!")
                cadastro_vazio = True
                tela_aberta = True
                break
                
            if agente.localizar_imagem("tela_inteira.png", confidence=0.7) or \
               agente.localizar_imagem("tela_acumuladores.png", confidence=0.7) or \
               agente.localizar_imagem("btn_cancelar2.png", confidence=0.7) or \
               agente.localizar_imagem("tela_cancelar2.png", confidence=0.7) or \
               agente.localizar_imagem("tela_novo.png", confidence=0.7) or \
               agente.localizar_imagem("btn_novo2.png", confidence=0.7):
                tela_aberta = True
                break
                
            time.sleep(1)
            
        if tela_aberta:
            if not cadastro_vazio:
                LOG.info(f"[VISÃO] View da tela detectada em {round(time.time() - tempo_espera, 1)}s!")
            break
        else:
            LOG.info("[AVISO] Tela não apareceu no tempo limite.")

    if not tela_aberta:
        raise RuntimeError("Falha ao abrir a tela de Acumuladores (nenhuma imagem detectada).")

    if cadastro_vazio:
        agente.pressionar("enter", 0.5)
        raise SemAcumuladoresException("Empresa não possui acumuladores (Aviso de Cadastro Vazio detectado).")

    LOG.info("[INFO] Esperando 5 segundos para o Domínio processar a tela completamente...")
    time.sleep(5.0)

    estado_tela = "DESCONHECIDO"
    LOG.info("[INFO] Verificando estado dos botões...")
    
    tempo_botoes = time.time()
    while (time.time() - tempo_botoes) < 15.0:
        if agente.localizar_imagem("tela_acumuladores.png", confidence=0.7) or \
           agente.localizar_imagem("tela_acumulador.png", confidence=0.7) or \
           agente.localizar_imagem("btn_cancelar2.png", confidence=0.7) or \
           agente.localizar_imagem("tela_cancelar2.png", confidence=0.7):
            estado_tela = "CANCELAR"
            break
            
        elif agente.localizar_imagem("tela_novo.png", confidence=0.7) or \
             agente.localizar_imagem("btn_novo2.png", confidence=0.7):
            estado_tela = "NOVO"
            break
            
        time.sleep(1.0)

    if estado_tela == "DESCONHECIDO":
        LOG.info("[OCR] Imagens falharam. Lendo o texto dentro das coordenadas exatas do botão (X:1270, Y:285)...")
        try:
            img_botao = pyautogui.screenshot(region=(1270, 285, 95, 40))
            texto_botao = pytesseract.image_to_string(img_botao, lang='por').lower()
            
            if "cancel" in texto_botao:
                estado_tela = "CANCELAR"
            elif "novo" in texto_botao:
                estado_tela = "NOVO"
        except Exception as e:
            LOG.info(f"[OCR] Erro ao ler o botão: {e}")

    if estado_tela == "CANCELAR":
        LOG.info(f"[VISÃO/OCR] Botão 'Cancelar' detectado!")
        LOG.info(f"[AÇÃO] Clicando no centro das coordenadas (X:{CFG.x_cancelar}, Y:{CFG.y_cancelar})")
        pyautogui.click(CFG.x_cancelar, CFG.y_cancelar)
        
        LOG.info("[INFO] Aguardando 5 segundos para verificar se a tela liberou (botão 'Novo' deve aparecer)...")
        time.sleep(5.0)
        
        achou_novo = False
        for _ in range(5):
            if agente.localizar_imagem("tela_novo.png", confidence=0.7) or agente.localizar_imagem("btn_novo2.png", confidence=0.7):
                achou_novo = True
                break
            time.sleep(1)
            
        if achou_novo:
            LOG.info("[VISÃO] Confirmação Visual: Botão 'Novo' apareceu após Cancelar! Tudo certo.")
        else:
            LOG.info("[FALLBACK] Botão 'Novo' não apareceu. Usando Opção Secundária (Point 871, 524 + 10 TABs).")
            pyautogui.click(871, 524)
            time.sleep(0.5)
            
            pausa_original = pyautogui.PAUSE
            pyautogui.PAUSE = 0.0
            pyautogui.press('tab', presses=10, interval=0.1)
            pyautogui.PAUSE = pausa_original
            
            agente.pressionar("enter", CFG.tempo_dominio_baixo)
            
    elif estado_tela == "NOVO":
        LOG.info("[VISÃO/OCR] Botão 'Novo' detectado logo de cara. Pulando Cancelar e indo para Listagem!")
        time.sleep(CFG.tempo_dominio_baixo)
        
    else:
        LOG.info("[AVISO FATAL] Não encontrou nem Novo nem Cancelar. Forçando Opção Secundária direto...")
        pyautogui.click(871, 524)
        time.sleep(0.5)
        
        pausa_original = pyautogui.PAUSE
        pyautogui.PAUSE = 0.0
        pyautogui.press('tab', presses=10, interval=0.1)
        pyautogui.PAUSE = pausa_original
        
        agente.pressionar("enter", CFG.tempo_dominio_baixo)

def gerar_relatorio_rpa(agente: AgenteDominio, codigo_empresa: int) -> None:
    LOG.info("[INFO] Iniciando geração de Relatório com atalhos puros...")
    
    LOG.info("[ATALHO] Pressionando (Alt + L) para Listagem...")
    agente.hotkey("alt", "l", seg=CFG.tempo_dominio_baixo)
        
    LOG.info("[ATALHO] Pressionando (Alt + R) para Relatório...")
    agente.hotkey("alt", "r", seg=CFG.tempo_dominio_baixo)

    LOG.info("[INFO] Aguardando tela de listagem abrir ('tela_listagem_aberta.png')...")
    tempo_espera = time.time()
    while (time.time() - tempo_espera) < CFG.tempo_dominio_max:
        if agente.localizar_imagem("tela_listagem_aberta.png", confidence=0.7):
            LOG.info("[VISÃO] View da Listagem Abriu!")
            break
        time.sleep(1)

    agente.pressionar("down", CFG.tempo_dominio_baixo)

    pausa_original = pyautogui.PAUSE
    pyautogui.PAUSE = 0.0
    pyautogui.press('tab', presses=20, interval=0.02)
    pyautogui.PAUSE = pausa_original
    
    agente.pressionar("enter", CFG.tempo_dominio)

    salvou = False
    max_tentativas = 3
    
    for tentativa in range(max_tentativas):
        LOG.info(f"--- Tentativa {tentativa+1}/{max_tentativas} de clicar no Excel ---")
        
        pos_excel = agente.localizar_imagem("tela_excel.png", confidence=0.7)
        if pos_excel:
            LOG.info(f"[VISÃO] Imagem 'tela_excel.png' encontrada! Clicando em X:{pos_excel.x} Y:{pos_excel.y}")
            pyautogui.click(pos_excel.x, pos_excel.y)
        else:
            LOG.info(f"[COORDENADAS] Imagem não achada. Clicando na opção Excel por Coordenada (X:{CFG.x_excel}, Y:{CFG.y_excel})")
            pyautogui.click(CFG.x_excel, CFG.y_excel)
            
        pyautogui.moveTo(CFG.x_voltar_dominio, CFG.y_voltar_dominio)
        time.sleep(1.0)
        
        LOG.info("[INFO] Aguardando a imagem 'tela_salvar_excel.png' aparecer...")
        tempo_salvar = time.time()
        pos_salvar = None
        
        while (time.time() - tempo_salvar) < 10.0: 
            pos_salvar = agente.localizar_imagem("tela_salvar_excel.png", confidence=0.8)
            if pos_salvar:
                LOG.info(f"[VISÃO] Janela 'Salvar' detectada em {round(time.time() - tempo_salvar, 1)}s!")
                break
            time.sleep(1)
            
        if pos_salvar:
            LOG.info("[INFO] Preenchendo o nome do arquivo...")
            pyautogui.click(CFG.x_nome_arquivo, CFG.y_nome_arquivo)
            time.sleep(0.5)
            agente.escrever(f" - {codigo_empresa}", 1.0)
            
            LOG.info("[VISÃO] Clicando no botão Salvar da janela do Windows.")
            pyautogui.click(CFG.x_salvar, CFG.y_salvar)
            salvou = True
            break
        else:
            LOG.info("[AVISO] Janela 'Salvar' não apareceu. O robô vai tentar clicar no ícone do Excel de novo!")

    if not salvou:
        LOG.info("[FALHA VISUAL] Janela de salvar não detectada após as tentativas. Executando salvamento às cegas (Fallback)!")
        pyautogui.click(CFG.x_nome_arquivo, CFG.y_nome_arquivo)
        time.sleep(0.5)
        agente.escrever(f" - {codigo_empresa}", 1.0)
        pyautogui.click(CFG.x_salvar, CFG.y_salvar)

    time.sleep(CFG.tempo_dominio)

    LOG.info("[INFO] Fechando Excel (Alt+F4) e voltando foco ao Domínio.")
    agente.hotkey("alt", "f4", seg=CFG.tempo_dominio_baixo)
    
    pyautogui.click(CFG.x_voltar_dominio, CFG.y_voltar_dominio)
    time.sleep(CFG.tempo_dominio_baixo)

    agente.pressionar("esc", CFG.tempo_dominio_baixo)

def extrair_relatorio_seguro(agente: AgenteDominio, codigo_empresa: int) -> Path:
    agente.fechar_sistema_para_loop(quantidade_esc=4, intervalo=0.3)
    entrar_em_acumuladores_por_atalho(agente)
    gerar_relatorio_rpa(agente, codigo_empresa)
    caminho_relatorio = aguardar_relatorio_empresa(codigo_empresa, timeout=15)
    return caminho_relatorio


# ============================================================
# INATIVAÇÃO COM MÚLTIPLAS VIGÊNCIAS (E LIMPEZA DE IMPOSTOS)
# ============================================================
def executar_inativacao_rpa(agente: AgenteDominio, codigo_acumulador: int, quantidade_vigencias: int) -> None:
    str_cod = str(codigo_acumulador)
    LOG.info(f"--- Iniciando inativação do acumulador {str_cod} ({quantidade_vigencias} vigência(s) encontrada(s)) ---")
    
    LOG.info("[ATALHO] Abrindo Listagem (Alt + L)...")
    agente.hotkey("alt", "l", seg=2.0)
    
    LOG.info(f"[INFO] Buscando o código {str_cod}...")
    agente.clicar_imagem_ou_coordenada("campo_busca.png", CFG.x_buscar_campo, CFG.y_buscar_campo, 1.0, descricao="Campo busca")
    agente.limpar_campo(1.0)
    agente.escrever(str_cod, 1.0)
    
    if str_cod == "1":
        LOG.info("[AVISO] Código 1 detectado. Selecionando o segundo item da lista...")
        agente.clicar_imagem_ou_coordenada("segundo_codigo_1.png", CFG.x_segundo_codigo_1, CFG.y_segundo_codigo_1, 1.0, descricao="Segundo 1")

    LOG.info("[ATALHO] Buscando acumulador (Alt + B)...")
    agente.hotkey("alt", "b", seg=CFG.tempo_dominio_max) 
    
    for i in range(quantidade_vigencias):
        if i > 0:
            LOG.info(f"[INFO] Processando a vigência {i+1} de {quantidade_vigencias} do acumulador {str_cod}...")
            LOG.info(f"[COORDENADAS] Voltando o foco para a Listagem (X:{CFG.x_foco_listagem}, Y:{CFG.y_foco_listagem})...")
            
            pyautogui.click(CFG.x_foco_listagem, CFG.y_foco_listagem)
            time.sleep(1.0)
            
            LOG.info("[TECLADO] Pressionando seta para baixo para selecionar a próxima vigência...")
            agente.pressionar("down", CFG.tempo_dominio) 
        
        # 1. CLICA NA SITUAÇÃO
        LOG.info("[INFO] Modificando situação para Inativo...")
        pyautogui.click(CFG.x_situacao, CFG.y_situacao)
        time.sleep(0.5)
        
        # 2. SETA PARA BAIXO
        agente.pressionar("down", 0.5)

        # ------------------------------------------------------------
        # TRATAMENTO DA MENSAGEM DE IMPOSTOS VINCULADOS
        # ------------------------------------------------------------
        LOG.info("[INFO] Verificando se há aviso de impostos vinculados...")
        tempo_aviso = time.time()
        tem_aviso = False
        
        while (time.time() - tempo_aviso) < 3.0:
            if agente.localizar_imagem(r"inativacao\tela_aviso_inativar1.png", confidence=0.7) or \
               agente.localizar_imagem("tela_aviso_inativar1.png", confidence=0.7):
                tem_aviso = True
                break
            time.sleep(0.5)
            
        if tem_aviso:
            LOG.info("[VISÃO] Aviso de impostos detectado! Fechando e limpando impostos...")
            agente.pressionar("enter", 1.0) # Aperta OK no Aviso
            
            # Clicar na aba Impostos
            pos_impostos = agente.localizar_imagem(r"inativacao\cabecalho2_impostos.png", confidence=0.7) or \
                           agente.localizar_imagem("cabecalho2_impostos.png", confidence=0.7)
            if pos_impostos:
                LOG.info("[VISÃO] Aba de Impostos encontrada. Clicando...")
                pyautogui.click(pos_impostos.x, pos_impostos.y)
            else:
                LOG.info(f"[FALLBACK] Clicando na aba Impostos por coordenada (X:{CFG.x_aba_impostos}, Y:{CFG.y_aba_impostos}).")
                pyautogui.click(CFG.x_aba_impostos, CFG.y_aba_impostos)
            
            time.sleep(1.0)
            
            # Excluir 5 vezes
            LOG.info(f"[AÇÃO] Clicando em Excluir 5 vezes (X:{CFG.x_excluir_imposto}, Y:{CFG.y_excluir_imposto})...")
            for _ in range(5):
                pyautogui.click(CFG.x_excluir_imposto, CFG.y_excluir_imposto)
                time.sleep(0.2)
                
            time.sleep(1.0)
            
            # COMO FOMOS LÁ EMBAIXO EXCLUIR, PRECISAMOS DEVOLVER O FOCO LÁ PRA CIMA ANTES DO TAB!
            LOG.info("[INFO] Devolvendo o foco para o campo Situação antes de dar o TAB...")
            pyautogui.click(CFG.x_situacao, CFG.y_situacao)
            time.sleep(0.5)
        # ------------------------------------------------------------
        
        # 3. UM TAB E DUAS SETAS PARA BAIXO (Exatamente como você pediu)
        LOG.info("[INFO] Pulando para a Data (1x TAB) e alterando (2x Seta Baixo)...")
        agente.pressionar("tab", 0.5)
        agente.pressionar("down", 0.2)
        agente.pressionar("down", 0.2)

        # 4. ALT + G (GRAVAR)
        if not CFG.dry_run:
            LOG.info("[ATALHO] Gravando alteração (Alt + G)...")
            agente.hotkey("alt", "g", seg=CFG.tempo_dominio_baixo)
            
            # Enter de segurança para fechar pop-ups de "Gravado com sucesso"
            agente.pressionar("enter", 1.0)
        else:
            LOG.info("[DRY RUN] Simulação ativada. Nenhuma gravação foi feita (Alt + G ignorado).")
            
    LOG.info(f"[SUCESSO] Acumulador {str_cod} processado com sucesso!")


def criar_log_empresa(codigo_empresa: int, nome_empresa: str, inativados: list, erros: list, status: str = "OK") -> None:
    caminho = CFG.pasta_inativacao / f"{codigo_empresa}.txt"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("===================================\n")
        f.write(f"Empresa: {codigo_empresa} - {nome_empresa}\n")
        f.write(f"Status Geral da Empresa: {status}\n")
        f.write("===================================\n")
        f.write("Acumuladores inativados:\n")
        for item in inativados: f.write(f"{item}\n")
        if not inativados: f.write("Nenhum\n")
        
        f.write("===================================\n")
        f.write("Erros:\n")
        for err in erros: f.write(f"{err}\n")
        if not erros: f.write("Nenhum\n")
        f.write("===================================\n")
        f.write("Fim do log.\n")

def registrar_execucao_csv(empresa: int, nome: str, acumulador: str, status: str, motivo: str) -> None:
    arquivo_existe = CFG.arquivo_execucao_csv.exists()
    campos =["timestamp", "empresa", "nome_empresa", "acumulador", "status", "motivo"]
    with open(CFG.arquivo_execucao_csv, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        if not arquivo_existe: writer.writeheader()
        writer.writerow({
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "empresa": empresa,
            "nome_empresa": nome,
            "acumulador": acumulador,
            "status": status,
            "motivo": motivo,
        })


# ============================================================
# PRÉ-CLASSIFICAÇÃO
# ============================================================
def classificar_empresas(df_empresas: pd.DataFrame) -> tuple[list, list]:
    precisa_extrair = []
    so_inativar     =[]

    for _, row in df_empresas.iterrows():
        codigo_empresa = int(row["CODIGO_EMPRESA"])
        nome_empresa   = row["NOME_EMPRESA"]

        if (CFG.pasta_inativacao / f"{codigo_empresa}.txt").exists():
            continue

        caminho = verificar_relatorio_existente(codigo_empresa)

        if caminho and arquivo_xls_valido(caminho):
            so_inativar.append((codigo_empresa, nome_empresa, caminho))
        else:
            if caminho:
                LOG.info(f"[AVISO] Relatório de {codigo_empresa} existe mas está corrompido — será re-extraído.")
                caminho.unlink()
            precisa_extrair.append((codigo_empresa, nome_empresa))

    return precisa_extrair, so_inativar


# ============================================================
# FLUXO PRINCIPAL
# ============================================================
def executar_fluxo_unificado():
    LOG.info("=== INICIANDO FLUXO UNIFICADO (ORDEM CRESCENTE) ===")

    validar_ambiente()

    df_empresas = carregar_empresas(CFG.planilha_empresas)
    LOG.info(f"[INFO] Foram carregadas {len(df_empresas)} empresas da planilha base.")

    if df_empresas.empty:
        LOG.info("[ERRO FATAL] Nenhuma empresa encontrada. Verifique a planilha base.")
        return

    lista_acumuladores_principais = carregar_acumuladores_principais(CFG.planilha_acumuladores)

    agente = criar_agente()
    LOG.info("Posicione a tela do Domínio Ativa. Iniciando em 5 segundos...")
    time.sleep(CFG.segundos_iniciais)

    for _, row in df_empresas.iterrows():
        codigo_empresa = int(row["CODIGO_EMPRESA"])
        nome_empresa   = row["NOME_EMPRESA"]
        
        arquivo_log = CFG.pasta_inativacao / f"{codigo_empresa}.txt"
        if arquivo_log.exists():
            LOG.info(f"--- [PULANDO] Empresa: {codigo_empresa} ({nome_empresa}) já tem LOG na pasta Inativação! ---")
            continue

        LOG.info(f"--- Processando Empresa: {codigo_empresa} ({nome_empresa}) ---")
        
        caminho_relatorio = verificar_relatorio_existente(codigo_empresa)
        precisa_extrair = False
        
        if caminho_relatorio and arquivo_xls_valido(caminho_relatorio):
            LOG.info("[INFO] Relatório válido encontrado na pasta. Pulando extração.")
        else:
            if caminho_relatorio:
                LOG.info(f"[AVISO] Relatório de {codigo_empresa} corrompido — será re-extraído.")
                caminho_relatorio.unlink()
            precisa_extrair = True

        inativados, erros = [],[]
        status_geral_log = "OK"

        try:
            executar_com_retry(lambda: agente.pressionar("f8", CFG.tempo_dominio_baixo), descricao="F8")
            executar_com_retry(lambda ce=codigo_empresa: agente.escrever(str(ce), CFG.tempo_dominio_baixo), descricao="Digitar Código")
            executar_com_retry(lambda: agente.pressionar("enter", CFG.tempo_dominio_baixo), descricao="Entrar na Empresa")

            if precisa_extrair:
                caminho_relatorio = executar_com_retry(
                    lambda ce=codigo_empresa: extrair_relatorio_seguro(agente, ce),
                    descricao=f"Extração — Empresa {codigo_empresa}"
                )
                if caminho_relatorio.suffix == ".xls" and not arquivo_xls_valido(caminho_relatorio):
                    raise RuntimeError("Arquivo extraído está corrompido.")

            acumuladores_empresa = processar_relatorio_extraido(caminho_relatorio)
            codigos_amarelos = {a["ACUMULADOR"] for a in lista_acumuladores_principais if a["AMARELO_PLANILHA"] == "SIM"}
            
            alvos_filtrados =[cod for cod in acumuladores_empresa if cod in codigos_amarelos]
            contagem_alvos = Counter(alvos_filtrados)

            LOG.info(f"[INFO] {len(contagem_alvos)} Acumuladores encontrados para inativar na empresa {codigo_empresa}")

            if contagem_alvos:
                if not precisa_extrair:
                    entrar_em_acumuladores_por_atalho(agente)

                for cod_alvo, quantidade in contagem_alvos.items():
                    try:
                        executar_com_retry(lambda ca=cod_alvo, qtd=quantidade: executar_inativacao_rpa(agente, ca, qtd), descricao=f"Inativar {cod_alvo}")
                        inativados.append(f"{cod_alvo} ({quantidade}x)")
                        registrar_execucao_csv(codigo_empresa, nome_empresa, str(cod_alvo), "INATIVADO", "Sucesso")
                    except Exception as e_item:
                        LOG.info(f"[ERRO] Falha ao inativar acumulador {cod_alvo}: {e_item}")
                        erros.append(f"Falha {cod_alvo}: {e_item}")
                        registrar_execucao_csv(codigo_empresa, nome_empresa, str(cod_alvo), "ERRO", str(e_item))
            else:
                status_geral_log = "PROCESSADA (SEM ALVOS)"
                registrar_execucao_csv(codigo_empresa, nome_empresa, "N/A", "PULADO", "Nenhum alvo encontrado")

        except SemAcumuladoresException as e_vazio:
            status_geral_log = "PROCESSADA (SEM ACUMULADORES CADASTRADOS)"
            LOG.info(f"[PULANDO] {e_vazio}")
            registrar_execucao_csv(codigo_empresa, nome_empresa, "N/A", "PULADO", "Empresa vazia")

        except Exception as e_geral:
            status_geral_log = "ERRO FATAL"
            LOG.info(f"[ERRO FATAL] Empresa {codigo_empresa}: {e_geral}")
            erros.append(f"Falha Crítica: {e_geral}")
            registrar_execucao_csv(codigo_empresa, nome_empresa, "N/A", "ERRO_EMPRESA", str(e_geral))
            try:
                pyautogui.screenshot(str(CFG.pasta_screenshots / f"ERRO_EMPRESA_{codigo_empresa}.png"))
            except Exception as e_ss:
                LOG.info(f"[AVISO] Screenshot não salvo: {e_ss}")

        finally:
            criar_log_empresa(codigo_empresa, nome_empresa, inativados, erros, status=status_geral_log)
            LOG.info("[INFO] Resetando Domínio (ESCs)...")
            agente.fechar_sistema_para_loop(CFG.esc_loop_qtd, CFG.esc_loop_intervalo)

    LOG.info("=== PROCESSO FINALIZADO ===")

if __name__ == "__main__":
    executar_fluxo_unificado()