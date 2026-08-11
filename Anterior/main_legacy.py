from __future__ import annotations

import csv
import re
import shutil
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pytesseract import Output

try:
    import pyautogui
except Exception:
    pyautogui = None

try:
    import win32com.client as win32
except Exception:
    win32 = None

from agente_dominio import AgenteConfig, AgenteDominio


# ============================================================
# EXCEÇÕES PERSONALIZADAS DO PROJETO
# ============================================================
class SemAcumuladoresException(Exception):
    pass


class ParadaManualException(Exception):
    pass


# ============================================================
# CONFIGURACAO UNIFICADA
# ============================================================
@dataclass(frozen=True)
class Config:
    base_dir: Path = Path.home() / "Desktop" / "PYTHON DOCS" / "P1 - Inativar Acumulador"

    planilha_empresas: Path = base_dir / "EMPRESA PARA INATIVAR ACUMULADOR.xlsx"
    planilha_acumuladores: Path = base_dir / "RELAÇÃO DE ACUMULADORES.xlsx"

    pasta_relatorios: Path = base_dir / "Relação Por empresa"
    pasta_logs: Path = base_dir / "Relatorio Final"
    pasta_screenshots: Path = pasta_logs / "screenshots"
    pasta_debug_agente: Path = pasta_logs / "debug_agente"
    pasta_inativacao: Path = base_dir / "Relatório Inativação"
    pasta_relatorios_finais: Path = base_dir / "Relação por empresa 3"
    pasta_imagens: Path = base_dir / "imagens_rpa"

    arquivo_execucao_csv: Path = pasta_logs / "execucao_consolidada.csv"
    arquivo_parada_manual: Path = pasta_logs / "PARAR.txt"

    # --- REGRAS DE NEGÓCIO ATUALIZADAS ---
    faixa_inativar_min: int = 1
    faixa_inativar_max: int = 211
    acumuladores_excluidos_inativacao: tuple[int, ...] = (2,)

    faixa_marcacao_min: int = 212
    faixa_marcacao_max: int = 1000

    data_inativacao_label: str = "042026"
    dry_run: bool = False

    caminho_tesseract: str = r"C:\Users\guilherme.rossi\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
    usar_ocr: bool = True

    pyautogui_pause: float = 0.5

    tempo_excel: float = 4.0
    tempo_dominio_baixo: float = 7.5
    tempo_dominio: float = 16.0
    tempo_dominio_max: float = 27.0

    retry_tentativas: int = 3
    retry_espera: float = 2.0

    esc_loop_qtd: int = 6
    esc_loop_intervalo: float = 0.3
    segundos_iniciais: int = 5

    x_cancelar: int = 1316
    y_cancelar: int = 304

    x_foco_listagem: int = 1527
    y_foco_listagem: int = 308

    x_excel: int = 29
    y_excel: int = 302

    x_nome_arquivo: int = 359
    y_nome_arquivo: int = 418

    x_salvar: int = 603
    y_salvar: int = 414

    x_voltar_dominio: int = 559
    y_voltar_dominio: int = 581

    x_situacao: int = 932
    y_situacao: int = 349

    x_inativo: int = 922
    y_inativo: int = 378

    x_data: int = 1192
    y_data: int = 356

    x_data_retry: int = 1182
    y_data_retry: int = 352

    x_gravar: int = 1326
    y_gravar: int = 370

    x_aba_impostos: int = 741
    y_aba_impostos: int = 462

    x_excluir_imposto: int = 1176
    y_excluir_imposto: int = 770

    regiao_situacao: tuple[int, int, int, int] = (892, 333, 90, 40)
    regiao_popup_central: tuple[int, int, int, int] = (600, 300, 720, 480)


CFG = Config()


class Logger:
    def __init__(self) -> None:
        CFG.pasta_logs.mkdir(parents=True, exist_ok=True)
        self.caminho = CFG.pasta_logs / "execucao_geral.log"

    def info(self, mensagem: str) -> None:
        linha = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {mensagem}"
        print(linha)
        with open(self.caminho, "a", encoding="utf-8") as f:
            f.write(linha + "\n")


LOG = Logger()


def validar_ambiente() -> None:
    if pyautogui is None:
        raise RuntimeError("Biblioteca 'pyautogui' não encontrada.")

    CFG.pasta_relatorios.mkdir(parents=True, exist_ok=True)
    CFG.pasta_screenshots.mkdir(parents=True, exist_ok=True)
    CFG.pasta_debug_agente.mkdir(parents=True, exist_ok=True)
    CFG.pasta_inativacao.mkdir(parents=True, exist_ok=True)
    CFG.pasta_relatorios_finais.mkdir(parents=True, exist_ok=True)
    CFG.pasta_imagens.mkdir(parents=True, exist_ok=True)

    if CFG.usar_ocr:
        caminho_tess = Path(CFG.caminho_tesseract)
        if caminho_tess.exists():
            pytesseract.pytesseract.tesseract_cmd = str(caminho_tess)
        else:
            LOG.info(f"[AVISO] Tesseract não encontrado em: {CFG.caminho_tesseract}. OCR será ignorado.")
            object.__setattr__(CFG, "usar_ocr", False)


def executar_com_retry(
    func: Callable,
    tentativas: int | None = None,
    espera: float | None = None,
    descricao: str = "Operação",
):
    if tentativas is None:
        tentativas = CFG.retry_tentativas
    if espera is None:
        espera = CFG.retry_espera

    ultimo_erro = None

    for tentativa in range(1, tentativas + 1):
        try:
            return func()
        except (SemAcumuladoresException, ParadaManualException):
            raise
        except Exception as exc:
            ultimo_erro = exc
            LOG.info(f"[RETRY] {descricao} falhou na tentativa {tentativa}/{tentativas}: {exc}")
            if tentativa < tentativas:
                time.sleep(espera)

    raise ultimo_erro


def normalizar_codigo(valor) -> Optional[int]:
    if pd.isna(valor): return None
    try:
        texto = str(valor).strip()
        if not texto: return None
        texto = texto.replace(".0", "")
        return int(float(texto))
    except Exception: return None


def cor_para_rgb(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid": return None
    if getattr(fill.start_color, "rgb", None): return fill.start_color.rgb.upper()
    if getattr(fill.start_color, "index", None): return str(fill.start_color.index).upper()
    return None


def criar_agente() -> AgenteDominio:
    return AgenteDominio(
        AgenteConfig(
            pasta_imagens=CFG.pasta_imagens,
            confidence=0.80,
            grayscale=False,
            pyautogui_pause=CFG.pyautogui_pause,
            tempo_padrao=CFG.tempo_dominio,
            tempo_max=CFG.tempo_dominio_max,
            usar_imagem=True,
            salvar_screenshots_debug=True,
            pasta_debug=CFG.pasta_debug_agente,
            failsafe=True,
        )
    )


def carregar_empresas(caminho: Path) -> pd.DataFrame:
    df = pd.read_excel(caminho)
    colunas = {str(c).lower().strip(): c for c in df.columns}
    coluna_codigo = next((colunas[n] for n in ("codigo", "código", "cod", "coluna a") if n in colunas), df.columns[0])
    coluna_empresa = next((colunas[n] for n in ("empresa", "nome", "razao social", "razão social") if n in colunas), None)

    df = df.copy()
    df["CODIGO_EMPRESA"] = df[coluna_codigo].apply(normalizar_codigo)
    df["NOME_EMPRESA"] = df[coluna_empresa].astype(str).str.strip() if coluna_empresa else ""
    df = df[df["CODIGO_EMPRESA"].notna()].copy()
    df["CODIGO_EMPRESA"] = df["CODIGO_EMPRESA"].astype(int)

    return df[["CODIGO_EMPRESA", "NOME_EMPRESA"]].drop_duplicates().sort_values("CODIGO_EMPRESA")


def carregar_acumuladores_principais(caminho_planilha: Path) -> list[dict]:
    wb = load_workbook(caminho_planilha, data_only=True)
    ws = wb.active

    if ws.max_row < 2 or ws.max_column < 1:
        raise ValueError("Planilha de acumuladores está vazia ou sem dados.")

    amarelos = {"FFFFFF00", "00FFFF00", "FFFF00", "FFFFFF99", "FFFFEB9C", "00FFEB9C"}
    registros: list[dict] =[]

    for row in ws.iter_rows(min_row=2):
        codigo = normalizar_codigo(row[0].value)
        if codigo is None:
            continue

        rgb = cor_para_rgb(row[0])
        marcado_amarelo = "SIM" if rgb in amarelos else "NAO"

        deve_inativar = (
            CFG.faixa_inativar_min <= codigo <= CFG.faixa_inativar_max
            and codigo not in CFG.acumuladores_excluidos_inativacao
            and marcado_amarelo == "SIM"
        )

        deve_marcar_analise = (
            CFG.faixa_marcacao_min <= codigo <= CFG.faixa_marcacao_max
        )

        registros.append({
            "ACUMULADOR": codigo,
            "AMARELO_PLANILHA": marcado_amarelo,
            "DEVE_INATIVAR": "SIM" if deve_inativar else "NAO",
            "DEVE_MARCAR_ANALISE": "SIM" if deve_marcar_analise else "NAO",
        })

    return registros


def verificar_relatorio_existente(codigo_empresa: int) -> Optional[Path]:
    nome_base = f"RELAÇÃO DE ACUMULADORES - {codigo_empresa}"
    caminho_xls = CFG.pasta_relatorios / f"{nome_base}.xls"
    caminho_xlsx = CFG.pasta_relatorios / f"{nome_base}.xlsx"

    if caminho_xlsx.exists(): return caminho_xlsx
    if caminho_xls.exists(): return caminho_xls
    return None


def arquivo_xls_valido(caminho: Path) -> bool:
    try:
        with open(caminho, "rb") as f: cabecalho = f.read(8)
        if caminho.suffix.lower() == ".xlsx": return cabecalho[:4] == b"PK\x03\x04"
        return cabecalho[:4] in (b"\xd0\xcf\x11\xe0", b"\x09\x08\x10\x00", b"\x01\x02\x06\x00")
    except Exception: return False


def aguardar_relatorio_empresa(codigo_empresa: int, timeout: int = 15, sufixo: str = "") -> Path:
    caminho_esperado = CFG.pasta_relatorios / f"RELAÇÃO DE ACUMULADORES - {codigo_empresa}{sufixo}.xls"
    for _ in range(timeout):
        if caminho_esperado.exists(): return caminho_esperado
        time.sleep(1)
    raise FileNotFoundError(f"O Excel '{caminho_esperado.name}' não apareceu na pasta.")


def converter_xls_para_xlsx_se_preciso(caminho_arquivo: Path) -> Path:
    if caminho_arquivo.suffix.lower() != ".xls": return caminho_arquivo
    destino = caminho_arquivo.with_suffix(".xlsx")
    if destino.exists() and arquivo_xls_valido(destino): return destino

    if win32 is None: raise RuntimeError("win32com não está disponível para converter .xls.")
    LOG.info("[INFO] Convertendo formato nativo do Domínio para .xlsx moderno via Excel...")

    excel = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(caminho_arquivo.resolve()))
        wb.SaveAs(str(destino.resolve()), FileFormat=51)
        wb.Close()
        excel.Quit()
        return destino
    except Exception as exc:
        if excel:
            try: excel.Quit()
            except Exception: pass
        raise RuntimeError(f"Falha ao converter arquivo usando Excel: {exc}")


def processar_relatorio_extraido(caminho_arquivo: Path) -> list[int]:
    caminho_arquivo = converter_xls_para_xlsx_se_preciso(caminho_arquivo)
    df = pd.read_excel(caminho_arquivo, header=None)
    df = df.iloc[6:].reset_index(drop=True)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    acumuladores: list[int] =[]
    if not df.empty:
        for valor in df.iloc[:, 0].tolist():
            codigo = normalizar_codigo(valor)
            if codigo is not None: acumuladores.append(codigo)
    return acumuladores


# ============================================================
# LÓGICA DE LOG BASEADA EM LINHA (MODO LOTE)
# ============================================================
def analisar_log_existente(codigo_empresa: int) -> tuple[bool, set[int], list[str]]:
    """ Lê o log e extrai as LINHAS (índices) já processadas para evitar retrabalho """
    caminho_log = CFG.pasta_inativacao / f"{codigo_empresa}.txt"
    if not caminho_log.exists(): return False, set(), []

    finalizada, status, tem_erros = False, "", False
    linhas_processadas = set()
    linhas_texto = []

    try:
        with open(caminho_log, "r", encoding="utf-8") as f:
            logs_txt = [l.strip() for l in f.readlines()]

        secao_inativados, secao_erros = False, False
        for l in logs_txt:
            if l.startswith("Status Geral da Empresa:"):
                status = l.split(":", 1)[1].strip()

            if l == "Acumuladores inativados:":
                secao_inativados, secao_erros = True, False
                continue

            if l == "Erros:":
                secao_erros, secao_inativados = True, False
                continue

            if l.startswith("==="):
                secao_inativados = secao_erros = False
                continue

            if secao_inativados and l and l != "Nenhum":
                linhas_texto.append(l)
                # Procura a tag [LINHA X] para registrar que ela já foi feita!
                match = re.search(r"\[LINHA (\d+)\]", l)
                if match:
                    linhas_processadas.add(int(match.group(1)))

            if secao_erros and l and l != "Nenhum":
                tem_erros = True

        statuses_finalizados = {
            "OK",
            "PROCESSADA (SEM ALVOS)",
            "PROCESSADA (SEM ACUMULADORES CADASTRADOS)",
        }
        finalizada = status in statuses_finalizados and not tem_erros
        return finalizada, linhas_processadas, linhas_texto

    except Exception as exc:
        LOG.info(f"[AVISO] Falha ao ler log existente da empresa {codigo_empresa}: {exc}")
        return False, set(), []


def criar_log_empresa(codigo_empresa: int, nome_empresa: str, inativados: list, erros: list, status: str = "OK") -> None:
    caminho = CFG.pasta_inativacao / f"{codigo_empresa}.txt"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("===================================\n")
        f.write(f"Empresa: {codigo_empresa} - {nome_empresa}\n")
        f.write(f"Status Geral da Empresa: {status}\n")
        f.write("===================================\n")
        f.write("Acumuladores inativados:\n")
        for item in inativados: f.write(f"{item}\n")
        if not inativados: f.write("Nenhum\n")
        f.write("===================================\n")
        f.write("Erros:\n")
        for err in erros: f.write(f"{err}\n")
        if not erros: f.write("Nenhum\n")
        f.write("===================================\n")
        f.write("Fim do log.\n")


def registrar_execucao_csv(empresa: int, nome: str, acumulador: str, status: str, motivo: str) -> None:
    arquivo_existe = CFG.arquivo_execucao_csv.exists()
    campos = ["timestamp", "empresa", "nome_empresa", "acumulador", "status", "motivo"]

    with open(CFG.arquivo_execucao_csv, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        if not arquivo_existe: writer.writeheader()
        writer.writerow({
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "empresa": empresa, "nome_empresa": nome,
            "acumulador": acumulador, "status": status, "motivo": motivo,
        })


def entrar_em_empresa(agente: AgenteDominio, codigo_empresa: int) -> None:
    if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")
    executar_com_retry(lambda: agente.pressionar("f8", CFG.tempo_dominio_baixo), descricao="F8")
    executar_com_retry(lambda: agente.escrever(str(codigo_empresa), CFG.tempo_dominio_baixo), descricao="Digitar Código")
    executar_com_retry(lambda: agente.pressionar("enter", CFG.tempo_dominio_baixo), descricao="Entrar na Empresa")


def entrar_em_acumuladores_por_atalho(agente: AgenteDominio) -> None:
    max_tentativas = 3
    tela_aberta = False
    cadastro_vazio = False

    for tentativa in range(1, max_tentativas + 1):
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        LOG.info(f"[INFO] Tentando entrar em Acumuladores (Alt+A+A) - Tentativa {tentativa}/{max_tentativas}")

        if tentativa > 1:
            estado_ja_aberto = (
                agente.localizar_imagem("tela_inteira.png", confidence=0.7)
                or agente.localizar_imagem("tela_acumuladores.png", confidence=0.7)
                or agente.localizar_imagem("tela_acumulador.png", confidence=0.7)
                or agente.localizar_imagem("btn_cancelar2.png", confidence=0.7)
                or agente.localizar_imagem("tela_cancelar2.png", confidence=0.7)
                or agente.localizar_imagem("tela_novo.png", confidence=0.7)
                or agente.localizar_imagem("btn_novo2.png", confidence=0.7)
            )

            if estado_ja_aberto:
                LOG.info("[VISÃO] Tela de acumuladores detectada com atraso.")
                break

            LOG.info("[INFO] Limpando tela com ESCs antes de tentar novamente.")
            agente.fechar_sistema_para_loop(quantidade_esc=3, intervalo=0.2)

        agente.sequencia_alt_aa(seg_entre_teclas=0.5, seg_final=1.0)
        agente.pressionar("enter", 0.5)

        LOG.info(f"[INFO] Aguardando tela do Domínio abrir até {CFG.tempo_dominio_max}s.")
        tempo_espera = time.time()
        
        while (time.time() - tempo_espera) < CFG.tempo_dominio_max:
            if agente.localizar_imagem("tela_aviso1.png", confidence=0.7) or agente.localizar_imagem("tela_aviso.png", confidence=0.7):
                LOG.info("[VISÃO] Aviso de 'cadastro vazio' detectado!")
                cadastro_vazio = True
                tela_aberta = True
                break
                
            if (agente.localizar_imagem("tela_inteira.png", confidence=0.7)
                or agente.localizar_imagem("tela_acumuladores.png", confidence=0.7)
                or agente.localizar_imagem("tela_acumulador.png", confidence=0.7)
                or agente.localizar_imagem("btn_cancelar2.png", confidence=0.7)
                or agente.localizar_imagem("tela_cancelar2.png", confidence=0.7)
                or agente.localizar_imagem("tela_novo.png", confidence=0.7)
                or agente.localizar_imagem("btn_novo2.png", confidence=0.7)):
                tela_aberta = True
                break

            time.sleep(1)

        if tela_aberta:
            if not cadastro_vazio:
                LOG.info(f"[VISÃO] View da tela detectada em {round(time.time() - tempo_espera, 1)}s!")
            break
        else:
            LOG.info("[AVISO] Tela não apareceu no tempo limite.")

    if not tela_aberta:
        raise RuntimeError("Falha ao abrir a tela de Acumuladores (nenhuma imagem detectada).")

    if cadastro_vazio:
        agente.pressionar("enter", 0.5)
        raise SemAcumuladoresException("Empresa não possui acumuladores cadastrados.")

    LOG.info("[INFO] Esperando 5 segundos para o Domínio processar a tela completamente.")
    time.sleep(5.0)

    # Estado Forçado: Se o Alt+C cancelar, destrava. Se o botão for "Novo", o Alt+C não faz mal.
    LOG.info("[ATALHO] Executando manobra de segurança: Pressionando Alt+C (Cancelar).")
    agente.hotkey("alt", "c", seg=CFG.tempo_dominio_baixo)


def abrir_listagem_relatorio(agente: AgenteDominio) -> None:
    LOG.info("[ATALHO] Pressionando Alt+L para Listagem.")
    agente.hotkey("alt", "l", seg=CFG.tempo_dominio_baixo)

    LOG.info("[ATALHO] Pressionando Alt+R para Relatório.")
    agente.hotkey("alt", "r", seg=CFG.tempo_dominio_baixo)

    LOG.info("[INFO] Aguardando tela de listagem/relatório abrir.")
    inicio = time.time()

    while time.time() - inicio < CFG.tempo_dominio_max:
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        if agente.localizar_imagem("tela_listagem_aberta.png", confidence=0.7):
            LOG.info("[VISÃO] Tela de Listagem/Relatório aberta.")
            return
        time.sleep(1)

    LOG.info("[FALLBACK] tela_listagem_aberta.png não detectada. Seguindo com teclas do roteiro.")


def gerar_relatorio_rpa(agente: AgenteDominio, codigo_empresa: int, sufixo: str = "") -> None:
    LOG.info("[INFO] Iniciando geração de relatório.")
    abrir_listagem_relatorio(agente)

    agente.pressionar("down", CFG.tempo_dominio_baixo)

    pausa_original = pyautogui.PAUSE
    pyautogui.PAUSE = 0.0
    pyautogui.press("tab", presses=20, interval=0.02)
    pyautogui.PAUSE = pausa_original

    agente.pressionar("enter", CFG.tempo_dominio)

    salvou = False
    max_tentativas = 3

    for tentativa in range(1, max_tentativas + 1):
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        LOG.info(f"--- Tentativa {tentativa}/{max_tentativas} de clicar na opção Excel ---")

        pos_excel = agente.localizar_imagem("tela_excel.png", confidence=0.7)
        if pos_excel:
            LOG.info("[VISÃO] Imagem 'tela_excel.png' encontrada.")
            pyautogui.click(pos_excel.x, pos_excel.y)
        else:
            LOG.info(f"[COORDENADAS] Clicando na opção Excel por coordenada (X:{CFG.x_excel}, Y:{CFG.y_excel}).")
            pyautogui.click(CFG.x_excel, CFG.y_excel)

        pyautogui.moveTo(CFG.x_voltar_dominio, CFG.y_voltar_dominio)
        time.sleep(1.0)

        LOG.info("[INFO] Aguardando janela 'Salvar Excel'.")
        inicio_salvar = time.time()
        pos_salvar = None

        while time.time() - inicio_salvar < 10.0:
            pos_salvar = agente.localizar_imagem("tela_salvar_excel.png", confidence=0.8)
            if pos_salvar:
                LOG.info("[VISÃO] Janela 'Salvar' detectada.")
                break
            time.sleep(1)

        if pos_salvar:
            LOG.info("[INFO] Preenchendo nome do arquivo.")
            pyautogui.click(CFG.x_nome_arquivo, CFG.y_nome_arquivo)
            time.sleep(0.5)
            agente.escrever(f" - {codigo_empresa}{sufixo}", 1.0)

            LOG.info("[VISÃO] Clicando no botão Salvar.")
            pyautogui.click(CFG.x_salvar, CFG.y_salvar)
            salvou = True
            break

        LOG.info("[AVISO] Janela 'Salvar' não apareceu. Tentando novamente.")

    if not salvou:
        LOG.info("[FALLBACK] Janela de salvar não detectada. Executando salvamento às cegas.")
        pyautogui.click(CFG.x_nome_arquivo, CFG.y_nome_arquivo)
        time.sleep(0.5)
        agente.escrever(f" - {codigo_empresa}{sufixo}", 1.0)
        pyautogui.click(CFG.x_salvar, CFG.y_salvar)

    time.sleep(CFG.tempo_dominio)

    LOG.info("[INFO] Fechando Excel com Alt+F4 e voltando foco ao Domínio.")
    agente.hotkey("alt", "f4", seg=CFG.tempo_dominio_baixo)
    pyautogui.click(CFG.x_voltar_dominio, CFG.y_voltar_dominio)
    time.sleep(CFG.tempo_dominio_baixo)
    agente.pressionar("esc", CFG.tempo_dominio_baixo)


def extrair_relatorio_seguro(agente: AgenteDominio, codigo_empresa: int, sufixo: str = "") -> Path:
    agente.fechar_sistema_para_loop(quantidade_esc=4, intervalo=0.3)
    entrar_em_acumuladores_por_atalho(agente)
    gerar_relatorio_rpa(agente, codigo_empresa, sufixo=sufixo)
    return aguardar_relatorio_empresa(codigo_empresa, timeout=15, sufixo=sufixo)


# ============================================================
# LÓGICA DE VERIFICAÇÃO DE INATIVIDADE E PROCESSAMENTO
# ============================================================
def verificar_se_ja_inativo(agente: AgenteDominio, str_cod: str) -> bool:
    achou_inativo = (
        agente.localizar_imagem("situacao_inativa.png", confidence=0.75, region=CFG.regiao_situacao) or 
        agente.localizar_imagem("situacao_inativa_azul.png", confidence=0.75, region=CFG.regiao_situacao)
    )

    if achou_inativo:
        agente.screenshot_debug(f"ja_inativo_detectado_{str_cod}")
        LOG.info(f"[VISÃO] Situação do acumulador {str_cod} detectada como INATIVA por imagem.")
        return True

    if CFG.usar_ocr:
        texto = agente.ocr_regiao(CFG.regiao_situacao, f"ocr_situacao_{str_cod}").lower()
        if texto:
            LOG.info(f"[OCR] Situação lida para acumulador {str_cod}: {texto}")

        if "inativ" in texto:
            agente.screenshot_debug(f"ja_inativo_detectado_ocr_{str_cod}")
            LOG.info(f"[OCR] Situação do acumulador {str_cod} detectada como INATIVA pelo OCR.")
            return True

    return False


def selecionar_situacao_inativo(agente: AgenteDominio, codigo_acumulador: str) -> None:
    LOG.info("[INFO] Selecionando situação Inativo.")
    pyautogui.click(CFG.x_situacao, CFG.y_situacao)
    time.sleep(0.5)
    agente.pressionar("down", 0.2)


def selecionar_data_inativacao_por_setas(agente: AgenteDominio, codigo_acumulador: str) -> None:
    LOG.info(f"[INFO] Selecionando data {CFG.data_inativacao_label} por 1x TAB.")
    # MUDANÇA: APENAS 1 TAB E ESCREVE DIRETAMENTE!
    agente.pressionar("tab", 0.5)
    agente.escrever(CFG.data_inativacao_label, 1.0)


def detectar_aviso_impostos(agente: AgenteDominio, timeout: float = 4.0) -> bool:
    LOG.info("[INFO] Verificando se há aviso de impostos vinculados.")
    inicio = time.time()

    while time.time() - inicio < timeout:
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        if (
            agente.localizar_imagem("inativacao\\tela_aviso_inativar1.png", confidence=0.7)
            or agente.localizar_imagem("tela_aviso_inativar1.png", confidence=0.7)
            or agente.localizar_imagem("tela_aviso.png", confidence=0.7)
        ):
            if not CFG.usar_ocr: return True

            texto_popup = agente.ocr_regiao(CFG.regiao_popup_central, "ocr_aviso_inativacao").lower()
            if texto_popup: LOG.info(f"[OCR] Texto do aviso: {texto_popup}")

            padroes_imposto = ["imposto", "vinculado", "vinculados", "acumulador"]

            if any(padrao in texto_popup for padrao in padroes_imposto): return True
            raise RuntimeError(f"Aviso desconhecido ao tentar inativar. OCR: {texto_popup}")

        if CFG.usar_ocr:
            texto_popup = agente.ocr_regiao(CFG.regiao_popup_central, "ocr_popup_central").lower()
            if texto_popup:
                if "imposto" in texto_popup or "vinculado" in texto_popup or "vinculados" in texto_popup:
                    LOG.info("[OCR] Popup de impostos vinculados detectado.")
                    return True

                if "aviso" in texto_popup and not any(p in texto_popup for p in ("imposto", "vinculado", "vinculados")):
                    raise RuntimeError(f"Aviso desconhecido ao tentar inativar. OCR: {texto_popup}")

        time.sleep(0.5)

    return False


def abrir_aba_impostos(agente: AgenteDominio) -> None:
    LOG.info("[INFO] Abrindo aba Impostos.")
    pos_impostos = (
        agente.localizar_imagem("inativacao\\aba_impostos.png", confidence=0.80)
        or agente.localizar_imagem("aba_impostos.png", confidence=0.80)
    )

    if pos_impostos:
        pyautogui.click(pos_impostos.x, pos_impostos.y)
        time.sleep(1.0)
        return

    LOG.info("[FALLBACK] Imagem 'aba_impostos.png' não encontrada. Usando coordenada da aba Impostos.")
    pyautogui.click(CFG.x_aba_impostos, CFG.y_aba_impostos)
    time.sleep(1.0)


def limpar_impostos_vinculados(agente: AgenteDominio) -> None:
    LOG.info("[VISÃO] Aviso de impostos detectado. Fechando aviso.")
    agente.pressionar("enter", 1.0)

    abrir_aba_impostos(agente)

    LOG.info(f"[AÇÃO] Clicando em Excluir imposto até 5 vezes (X:{CFG.x_excluir_imposto}, Y:{CFG.y_excluir_imposto}).")
    for _ in range(5):
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")
        
        pos_excluir = agente.localizar_imagem("btn_excluir.png", confidence=0.8)
        if pos_excluir:
            pyautogui.click(pos_excluir.x, pos_excluir.y)
        else:
            pyautogui.click(CFG.x_excluir_imposto, CFG.y_excluir_imposto)
            
        time.sleep(0.2)

    time.sleep(1.0)
    
    LOG.info("[INFO] Devolvendo o foco para o campo Situação...")
    pyautogui.click(CFG.x_situacao, CFG.y_situacao)
    time.sleep(0.5)


def gravar_inativacao(agente: AgenteDominio, codigo_acumulador: str) -> None:
    if CFG.dry_run:
        LOG.info("[DRY RUN] Simulação ativada. Nenhuma gravação foi feita.")
        return

    LOG.info("[ATALHO] Gravando alteração com Alt+G.")
    agente.hotkey("alt", "g", seg=2.0)
    
    LOG.info("[INFO] Verificando se o sistema rejeitou a data ou gerou pop-ups...")
    tempo_verificacao = time.time()
    
    while (time.time() - tempo_verificacao) < 5.0:
        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        # ERRO DE DATA INVÁLIDA
        if agente.localizar_imagem("tela_data_invalida.png", confidence=0.7):
            LOG.info("[ERRO FATAL] Pop-up 'Data de inativação inválida' detectado! Fechando e abortando este acumulador.")
            agente.pressionar("enter", 0.5) 
            agente.pressionar("esc", 0.5)   
            raise RuntimeError("Falha ao gravar: O sistema recusou a data de inativação (Data Inválida).")

        # ERRO DE IMPOSTO NÃO CADASTRADO NA VIGÊNCIA
        texto_popup = ""
        if CFG.usar_ocr:
            texto_popup = agente.ocr_regiao(CFG.regiao_popup_central, "ocr_popup_gravar").lower()
        
        if agente.localizar_imagem("aviso_imposto.png", confidence=0.7) or \
           "não está cadastrado" in texto_popup or "cadastrado nos parâmetros" in texto_popup:
            
            LOG.info("[VISÃO/OCR] Aviso de 'Imposto não cadastrado' detectado após tentar gravar!")
            LOG.info("[AÇÃO] Fechando pop-up de aviso com Enter...")
            agente.pressionar("enter", 1.0)
            
            LOG.info(f"[AÇÃO] Clicando na aba Impostos (X:{CFG.x_aba_impostos}, Y:{CFG.y_aba_impostos})...")
            pyautogui.click(CFG.x_aba_impostos, CFG.y_aba_impostos)
            time.sleep(1.0)
            
            LOG.info("[VISÃO] Procurando o botão 'Excluir'...")
            for _ in range(5):
                pos_excluir = agente.localizar_imagem("btn_excluir.png", confidence=0.8)
                if pos_excluir:
                    pyautogui.click(pos_excluir.x, pos_excluir.y)
                else:
                    pyautogui.click(CFG.x_excluir_imposto, CFG.y_excluir_imposto)
                time.sleep(0.2)
            
            time.sleep(1.0)
            
            LOG.info("[INFO] Devolvendo o foco para o campo Situação e tentando Gravar novamente...")
            pyautogui.click(CFG.x_situacao, CFG.y_situacao)
            time.sleep(0.5)
            
            LOG.info("[ATALHO] Gravando alteração com Alt+G (Segunda tentativa)...")
            agente.hotkey("alt", "g", seg=2.0)
            
            tempo_verificacao = time.time()
            continue

        time.sleep(0.5)

    LOG.info("[INFO] Gravação aparentemente aceita. Pressionando Enter de segurança.")
    agente.pressionar("enter", 1.0)


def processar_vigencia_atual(agente: AgenteDominio, str_cod: str) -> bool:
    LOG.info("[INFO] Verificando se a vigência atual já está Inativa...")

    if verificar_se_ja_inativo(agente, str_cod):
        return False

    LOG.info(f"[INFO] Acumulador {str_cod} está ATIVO. Iniciando inativação.")

    selecionar_situacao_inativo(agente, str_cod)
    tem_aviso = detectar_aviso_impostos(agente)

    if tem_aviso:
        limpar_impostos_vinculados(agente)
        LOG.info("[INFO] Após limpar impostos, selecionando Inativo novamente.")
        
        pyautogui.click(CFG.x_situacao, CFG.y_situacao)
        time.sleep(0.5)
        
        selecionar_situacao_inativo(agente, str_cod)

    selecionar_data_inativacao_por_setas(agente, str_cod)
    gravar_inativacao(agente, str_cod)

    return True

# ============================================================
# LÓGICA DE LOTE (BATCH MODE) 
# ============================================================
def executar_inativacao_em_lote(
    agente: AgenteDominio, 
    codigo_empresa: int,
    nome_empresa: str,
    acumuladores_empresa: list[int], 
    codigos_alvo: set[int],
    linhas_ja_processadas: set[int],
    inativados: list[str],
    erros: list[str]
) -> None:

    LOG.info("[ATALHO] Abrindo Listagem (Alt + L)...")
    agente.hotkey("alt", "l", seg=2.0)
    
    LOG.info("[ATALHO] Buscando TODOS os acumuladores da empresa (Alt + B)...")
    agente.hotkey("alt", "b", seg=CFG.tempo_dominio_max) 
    
    LOG.info(f"[COORDENADAS] Dando foco na Listagem (X:{CFG.x_foco_listagem}, Y:{CFG.y_foco_listagem}).")
    pyautogui.click(CFG.x_foco_listagem, CFG.y_foco_listagem)
    time.sleep(1.0)
    
    for i, cod in enumerate(acumuladores_empresa):
        agente.verificar_parada_manual()

        deve_inativar = (
            cod in codigos_alvo
            and CFG.faixa_inativar_min <= cod <= CFG.faixa_inativar_max
            and cod not in CFG.acumuladores_excluidos_inativacao
        )
        
        ja_processado_antes = i in linhas_ja_processadas

        if deve_inativar and not ja_processado_antes:
            LOG.info(f"--- [ALVO ENCONTRADO] Linha {i+1}/{len(acumuladores_empresa)} - Acumulador {cod} ---")
            
            try:
                processou = processar_vigencia_atual(agente, str(cod))
                
                if processou:
                    LOG.info(f"[OK] Acumulador {cod} (Linha {i+1}) foi inativado.")
                    inativados.append(f"[LINHA {i}] Acumulador {cod} - Inativado")
                    registrar_execucao_csv(codigo_empresa, nome_empresa, str(cod), "INATIVADO", f"Linha {i+1} Inativada")
                else:
                    LOG.info(f"[PULADO] Acumulador {cod} (Linha {i+1}) já estava inativo.")
                    inativados.append(f"[LINHA {i}] Acumulador {cod} - Já estava inativo")
                    registrar_execucao_csv(codigo_empresa, nome_empresa, str(cod), "PULADO", f"Linha {i+1} Já inativa")
                    
                linhas_ja_processadas.add(i)
                    
            except Exception as e_item:
                LOG.info(f"[ERRO] Falha ao inativar acumulador {cod}: {e_item}")
                erros.append(f"Falha [LINHA {i}] Acumulador {cod}: {e_item}")
                registrar_execucao_csv(codigo_empresa, nome_empresa, str(cod), "ERRO", str(e_item))
                agente.screenshot_debug(f"erro_lote_{codigo_empresa}_{cod}_linha_{i}")
                raise RuntimeError(f"Quebrando fluxo da empresa devido a erro na linha {i+1} (Acumulador {cod}): {e_item}")
                
        else:
            if i % 20 == 0: # Avisa no log a cada 20 linhas puladas pra não achar que travou
                LOG.info(f"[INFO] Percorrendo lista com Seta para Baixo... (Linha {i+1}/{len(acumuladores_empresa)})")

        if i < len(acumuladores_empresa) - 1:
            if deve_inativar and not ja_processado_antes:
                # Devolve foco para a listagem antes de descer se tiver mexido nos botões
                pyautogui.click(CFG.x_foco_listagem, CFG.y_foco_listagem)
                time.sleep(0.5)
            
            agente.pressionar("down", 0.2)


def colorir_planilha_verificacao(caminho_planilha: Path, codigos_inativados: set[int]) -> None:
    try:
        from openpyxl.styles import PatternFill
        wb = load_workbook(caminho_planilha)
        ws = wb.active
        
        cor_verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cor_amarela = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in ws.iter_rows(min_row=1): 
            celula_codigo = row[0]
            codigo = normalizar_codigo(celula_codigo.value)
            
            if codigo is None:
                continue
                
            if codigo in codigos_inativados:
                for cell in row: cell.fill = cor_verde
            elif CFG.faixa_marcacao_min <= codigo <= CFG.faixa_marcacao_max:
                for cell in row: cell.fill = cor_amarela
                        
        wb.save(caminho_planilha)
        LOG.info(f"[INFO] Planilha de verificação colorida: verde=inativados, amarelo=possíveis/análise.")
    except Exception as e:
        LOG.info(f"[AVISO] Falha não crítica ao tentar colorir a planilha de verificação: {e}")


def mover_relatorio_verificacao(caminho_verificacao: Path) -> Path:
    destino = CFG.pasta_relatorios_finais / caminho_verificacao.name
    if destino.exists(): destino.unlink()
    shutil.move(str(caminho_verificacao), str(destino))
    return destino

def extrair_verificacao_final(agente: AgenteDominio, codigo_empresa: int, inativados_log: list[str]) -> Path:
    LOG.info("[INFO] Extraindo relatório final de verificação.")
    caminho_verificacao = executar_com_retry(
        lambda ce=codigo_empresa: extrair_relatorio_seguro(agente, ce, sufixo=" - VERIFICACAO"),
        descricao=f"Extração de Verificação — Empresa {codigo_empresa}",
    )
    
    caminho_xlsx = converter_xls_para_xlsx_se_preciso(caminho_verificacao)
    
    if caminho_verificacao.exists() and caminho_verificacao.suffix == ".xls":
        try: caminho_verificacao.unlink()
        except Exception: pass
        
    destino = mover_relatorio_verificacao(caminho_xlsx)
    LOG.info(f"[SUCESSO] Relatório de verificação movido para: {destino}")
    
    # Extrai o número do acumulador das mensagens do Log para pintar na planilha
    codigos_processados = set()
    for item in inativados_log:
        match = re.search(r"Acumulador (\d+)", item)
        if match:
            codigos_processados.add(int(match.group(1)))
            
    colorir_planilha_verificacao(destino, codigos_processados)
    return destino


def calcular_alvos_empresa(
    acumuladores_empresa: list[int],
    lista_acumuladores_principais: list[dict],
    linhas_ja_processadas: set[int],
) -> Counter:
    codigos_alvo = {
        a["ACUMULADOR"]
        for a in lista_acumuladores_principais
        if a["DEVE_INATIVAR"] == "SIM"
    }

    alvos_filtrados = [
        cod
        for i, cod in enumerate(acumuladores_empresa)
        if cod in codigos_alvo
        and CFG.faixa_inativar_min <= cod <= CFG.faixa_inativar_max
        and cod not in CFG.acumuladores_excluidos_inativacao
        and i not in linhas_ja_processadas
    ]

    return Counter(alvos_filtrados)


def executar_fluxo_unificado() -> None:
    CFG.pasta_logs.mkdir(parents=True, exist_ok=True)
    LOG.info("=== INICIANDO FLUXO UNIFICADO (ORDEM CRESCENTE + BATCH MODE) ===")

    validar_ambiente()

    df_empresas = carregar_empresas(CFG.planilha_empresas)
    LOG.info(f"[INFO] Foram carregadas {len(df_empresas)} empresas da planilha base.")

    if df_empresas.empty:
        LOG.info("[ERRO FATAL] Nenhuma empresa encontrada. Verifique a planilha base.")
        return

    lista_acumuladores_principais = carregar_acumuladores_principais(CFG.planilha_acumuladores)

    agente = criar_agente()

    LOG.info(f"Posicione a tela do Domínio ativa. Iniciando em {CFG.segundos_iniciais} segundos.")
    time.sleep(CFG.segundos_iniciais)

    for _, row in df_empresas.iterrows():
        codigo_empresa = int(row["CODIGO_EMPRESA"])
        nome_empresa = str(row["NOME_EMPRESA"])

        if CFG.arquivo_parada_manual.exists(): raise ParadaManualException("Parada manual acionada.")

        finalizada, linhas_ja_processadas, linhas_texto_inativados = analisar_log_existente(codigo_empresa)

        if finalizada:
            LOG.info(f"--- [PULANDO] Empresa {codigo_empresa} ({nome_empresa}) já concluída com sucesso. ---")
            continue

        LOG.info(f"--- Processando Empresa: {codigo_empresa} ({nome_empresa}) ---")

        if linhas_ja_processadas:
            LOG.info(f"[INFO] Retomando empresa {codigo_empresa}. Linhas já processadas/puladas no Log: {sorted(list(linhas_ja_processadas))}")

        inativados = list(linhas_texto_inativados)
        erros: list[str] =[]
        status_geral_log = "OK"

        caminho_relatorio = verificar_relatorio_existente(codigo_empresa)
        precisa_extrair = False

        if caminho_relatorio and arquivo_xls_valido(caminho_relatorio):
            LOG.info("[INFO] Relatório válido encontrado na pasta. Pulando extração inicial.")
        else:
            if caminho_relatorio:
                LOG.info(f"[AVISO] Relatório de {codigo_empresa} corrompido. Será reextraído.")
                caminho_relatorio.unlink(missing_ok=True)
            precisa_extrair = True

        try:
            entrar_em_empresa(agente, codigo_empresa)

            if precisa_extrair:
                caminho_relatorio = executar_com_retry(
                    lambda ce=codigo_empresa: extrair_relatorio_seguro(agente, ce),
                    descricao=f"Extração — Empresa {codigo_empresa}",
                )

                if caminho_relatorio.suffix.lower() == ".xls" and not arquivo_xls_valido(caminho_relatorio):
                    raise RuntimeError("Arquivo extraído está corrompido.")

            acumuladores_empresa = processar_relatorio_extraido(caminho_relatorio)

            contagem_alvos = calcular_alvos_empresa(
                acumuladores_empresa=acumuladores_empresa,
                lista_acumuladores_principais=lista_acumuladores_principais,
                linhas_ja_processadas=linhas_ja_processadas,
            )

            LOG.info(f"[INFO] Restam {len(list(contagem_alvos.elements()))} linhas-alvo para inativar na empresa {codigo_empresa}.")

            if len(list(contagem_alvos.elements())) > 0:
                if not precisa_extrair:
                    entrar_em_acumuladores_por_atalho(agente)

                executar_inativacao_em_lote(
                    agente, 
                    codigo_empresa, 
                    nome_empresa, 
                    acumuladores_empresa, 
                    {a["ACUMULADOR"] for a in lista_acumuladores_principais if a["DEVE_INATIVAR"] == "SIM"}, 
                    linhas_ja_processadas, 
                    inativados, 
                    erros
                )

                try:
                    extrair_verificacao_final(agente, codigo_empresa, inativados)
                except Exception as exc_verif:
                    LOG.info(f"[ERRO] Falha ao extrair relatório de verificação: {exc_verif}")
                    erros.append(f"Verificação Falhou: {exc_verif}")

            else:
                if linhas_ja_processadas:
                    status_geral_log = "OK"
                else:
                    status_geral_log = "PROCESSADA (SEM ALVOS)"
                    registrar_execucao_csv(
                        codigo_empresa,
                        nome_empresa,
                        "N/A",
                        "PULADO",
                        "Nenhum alvo de inativação encontrado entre 1 e 211, exceto acumulador 2",
                    )

        except SemAcumuladoresException as exc_vazio:
            status_geral_log = "PROCESSADA (SEM ACUMULADORES CADASTRADOS)"
            LOG.info(f"[PULANDO] {exc_vazio}")
            registrar_execucao_csv(
                codigo_empresa,
                nome_empresa,
                "N/A",
                "PULADO",
                "Empresa vazia",
            )

        except ParadaManualException as exc_parada:
            status_geral_log = "INTERROMPIDO MANUALMENTE"
            LOG.info(f"[PARADA] {exc_parada}")
            erros.append(str(exc_parada))
            criar_log_empresa(codigo_empresa, nome_empresa, inativados, erros, status=status_geral_log)
            raise

        except Exception as exc_geral:
            status_geral_log = "ERRO FATAL"
            LOG.info(f"[ERRO FATAL] Empresa {codigo_empresa}: {exc_geral}")
            erros.append(f"Falha Crítica: {exc_geral}")
            registrar_execucao_csv(
                codigo_empresa,
                nome_empresa,
                "N/A",
                "ERRO_EMPRESA",
                str(exc_geral),
            )

            try:
                pyautogui.screenshot(str(CFG.pasta_screenshots / f"ERRO_EMPRESA_{codigo_empresa}.png"))
            except Exception as exc_ss:
                LOG.info(f"[AVISO] Screenshot não salvo: {exc_ss}")

        finally:
            if erros and status_geral_log == "OK":
                status_geral_log = "ERRO PARCIAL"

            criar_log_empresa(codigo_empresa, nome_empresa, inativados, erros, status=status_geral_log)

            LOG.info("[INFO] Resetando Domínio com ESCs.")
            try:
                agente.fechar_sistema_para_loop(CFG.esc_loop_qtd, CFG.esc_loop_intervalo)
            except Exception as exc_reset:
                LOG.info(f"[AVISO] Falha ao resetar Domínio: {exc_reset}")

    LOG.info("=== PROCESSO FINALIZADO ===")


if __name__ == "__main__":
    executar_fluxo_unificado()