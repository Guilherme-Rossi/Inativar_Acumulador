from __future__ import annotations

import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    import pyautogui
except Exception:
    pyautogui = None

from agente_dominio import AgenteConfig, AgenteDominio


# ============================================================
# CONFIGURACAO
# ============================================================
@dataclass(frozen=True)
class Config:
    base_dir: Path = Path(r"C:\Users\guilherme.rossi\Desktop\PYTHON DOCS\Projects Agrelli\P1 - Inativar Acumulador")
    planilha_empresas: Path = base_dir / "EMPRESA PARA INATIVAR ACUMULADOR.xlsx"
    planilha_acumuladores: Path = base_dir / "RELAÇÃO DE ACUMULADORES.xlsx"
    pasta_relatorios: Path = base_dir / "Relação Por empresa"
    pasta_logs: Path = base_dir / "Relatorio Final"
    arquivo_validados: Path = base_dir / "VALIDADOS.xlsx"
    arquivo_auditoria_csv: Path = base_dir / "AUDITORIA_VALIDACAO.csv"
    pasta_imagens: Path = base_dir / "imagens_rpa"
    pasta_debug_agente: Path = pasta_logs / "debug_agente"

    faixa_min: int = 212
    faixa_max: int = 1000

    coletar_relatorios_rpa: bool = True
    
    # --- OTIMIZACOES DE TEMPO ---
    pyautogui_pause: float = 0.1

    tempo_dominio_baixo: float = 2.0
    tempo_dominio: float = 10.5
    tempo_dominio_max: float = 13.5

    retry_tentativas: int = 3
    retry_espera: float = 2.0

    esc_loop_qtd: int = 6
    esc_loop_intervalo: float = 0.1
    segundos_iniciais: int = 5
    # ----------------------------

    cor_inativar_rgb: str = "92D050"
    cor_possivel_rgb: str = "FFF2CC"
    cor_nao_existe_rgb: str = "F4CCCC"

    # coordenadas mantidas como principal agora que a imagem esta desligada
    x_listagem: int = 1310
    y_listagem: int = 406
    x_tirar_relatorio: int = 1407
    y_tirar_relatorio: int = 706
    x_opcoes_relatorio: int = 822
    y_opcoes_relatorio: int = 351
    x_trocar_opcao: int = 842
    y_trocar_opcao: int = 379
    x_gerar: int = 1396
    y_gerar: int = 338
    x_excel: int = 36
    y_excel: int = 307
    x_nome_arquivo: int = 359
    y_nome_arquivo: int = 418
    x_salvar: int = 593
    y_salvar: int = 412
    
    # --- NOVAS COORDENADAS ---
    x_voltar_dominio: int = 559
    y_voltar_dominio: int = 581


CFG = Config()


def garantir_pastas() -> None:
    CFG.pasta_relatorios.mkdir(parents=True, exist_ok=True)
    CFG.pasta_logs.mkdir(parents=True, exist_ok=True)
    CFG.pasta_imagens.mkdir(parents=True, exist_ok=True)
    CFG.pasta_debug_agente.mkdir(parents=True, exist_ok=True)


class Auditor:
    def __init__(self) -> None:
        garantir_pastas()
        self.caminho = CFG.pasta_logs / "etapa1_validacao.log"

    def info(self, mensagem: str) -> None:
        linha = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {mensagem}"
        print(linha)
        with open(self.caminho, "a", encoding="utf-8") as f:
            f.write(linha + "\n")


LOG = Auditor()


def executar_com_retry(func, tentativas: int = None, espera: float = None, descricao: str = "Operacao"):
    if tentativas is None:
        tentativas = CFG.retry_tentativas
    if espera is None:
        espera = CFG.retry_espera

    ultimo_erro = None
    for tentativa in range(1, tentativas + 1):
        try:
            return func()
        except Exception as exc:
            ultimo_erro = exc
            LOG.info(f"[RETRY] {descricao} falhou na tentativa {tentativa}/{tentativas}: {exc}")
            if tentativa < tentativas:
                time.sleep(espera)
    raise ultimo_erro


def normalizar_codigo(valor) -> Optional[int]:
    if pd.isna(valor):
        return None
    try:
        texto = str(valor).strip()
        if not texto:
            return None
        texto = texto.replace(".0", "")
        return int(float(texto))
    except Exception:
        return None


def criar_agente() -> AgenteDominio:
    if pyautogui is None:
        raise RuntimeError("pyautogui nao esta instalado.")

    return AgenteDominio(
        AgenteConfig(
            pasta_imagens=CFG.pasta_imagens,
            confidence=0.80,
            grayscale=True,
            pyautogui_pause=CFG.pyautogui_pause,
            tempo_padrao=CFG.tempo_dominio,
            tempo_max=CFG.tempo_dominio_max,
            usar_imagem=False, # <--- IMAGEM DESATIVADA PARA MAXIMA VELOCIDADE
            usar_acessibilidade=False,
            salvar_screenshots_debug=True,
            pasta_debug=CFG.pasta_debug_agente,
        )
    )


# ============================================================
# LEITURA DAS PLANILHAS BASE
# ============================================================
def carregar_empresas(caminho: Path) -> pd.DataFrame:
    df = pd.read_excel(caminho)
    colunas = {c.lower().strip(): c for c in df.columns}

    coluna_codigo = None
    for nome in ("codigo", "código", "cod", "coluna a"):
        if nome in colunas:
            coluna_codigo = colunas[nome]
            break
    if coluna_codigo is None:
        coluna_codigo = df.columns[0]

    coluna_empresa = None
    for nome in ("empresa", "nome", "razao social", "razão social"):
        if nome in colunas:
            coluna_empresa = colunas[nome]
            break

    df = df.copy()
    df["CODIGO_EMPRESA"] = df[coluna_codigo].apply(normalizar_codigo)
    if coluna_empresa:
        df["NOME_EMPRESA"] = df[coluna_empresa].astype(str).str.strip()
    else:
        df["NOME_EMPRESA"] = ""

    df = df[df["CODIGO_EMPRESA"].notna()].copy()
    df["CODIGO_EMPRESA"] = df["CODIGO_EMPRESA"].astype(int)
    df = df[["CODIGO_EMPRESA", "NOME_EMPRESA"]].drop_duplicates().sort_values("CODIGO_EMPRESA")
    return df


def cor_para_rgb(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None

    color = fill.start_color
    rgb = getattr(color, "rgb", None)
    if rgb:
        return rgb.upper()

    indexed = getattr(color, "index", None)
    if indexed:
        return str(indexed).upper()
    return None


def carregar_acumuladores_principais(caminho_planilha: Path) -> pd.DataFrame:
    wb = load_workbook(caminho_planilha, data_only=True)
    ws = wb.active

    amarelos = {
        "FFFFFF00", "00FFFF00", "FFFF00", "FFFFFF99", "FFFFEB9C", "00FFEB9C"
    }

    registros =[]
    for row in ws.iter_rows(min_row=2):
        cell_codigo = row[0]
        codigo = normalizar_codigo(cell_codigo.value)
        if codigo is None:
            continue

        linha = cell_codigo.row
        nome = ws.cell(row=linha, column=2).value
        situacao = ws.cell(row=linha, column=6).value if ws.max_column >= 6 else None
        data_inatividade = ws.cell(row=linha, column=7).value if ws.max_column >= 7 else None
        rgb = cor_para_rgb(cell_codigo)
        marcado_amarelo = "SIM" if rgb in amarelos else "NAO"
        dentro_faixa = "SIM" if CFG.faixa_min <= codigo <= CFG.faixa_max else "NAO"

        registros.append({
            "ACUMULADOR": codigo,
            "NOME_ACUMULADOR": str(nome).strip() if nome is not None else "",
            "SITUACAO_PLANILHA": str(situacao).strip() if situacao is not None else "",
            "DATA_INATIVIDADE_PLANILHA": data_inatividade,
            "AMARELO_PLANILHA": marcado_amarelo,
            "DENTRO_FAIXA_212_1000": dentro_faixa,
            "COR_RGB": rgb or "",
        })

    df = pd.DataFrame(registros)
    if df.empty:
        raise ValueError("Nenhum acumulador foi lido na planilha principal.")

    df = df.drop_duplicates(subset=["ACUMULADOR"]).sort_values("ACUMULADOR")
    return df


# ============================================================
# RELATORIOS DAS EMPRESAS
# ============================================================
def converter_xls_para_xlsx(pasta: Path) -> None:
    for arquivo in pasta.glob("*.xls"):
        try:
            destino = arquivo.with_suffix(".xlsx")
            if destino.exists() and destino.stat().st_mtime >= arquivo.stat().st_mtime:
                LOG.info(f"[SKIP] Conversao nao necessaria: {arquivo.name}")
                continue
            df = pd.read_excel(arquivo, header=None)
            df.to_excel(destino, index=False, header=False)
            LOG.info(f"[OK] Convertido: {arquivo.name} -> {destino.name}")
        except Exception as exc:
            LOG.info(f"[ERRO] Falha ao converter {arquivo.name}: {exc}")


def localizar_relatorio_empresa(codigo_empresa: int, pasta: Path) -> Optional[Path]:
    candidatos = list(pasta.glob(f"*{codigo_empresa}*.xlsx"))
    candidatos =[p for p in candidatos if p.name.lower() != CFG.arquivo_validados.name.lower()]
    if not candidatos:
        return None
    candidatos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidatos[0]


def limpar_planilha_empresa(caminho_xlsx: Path) -> pd.DataFrame:
    df = pd.read_excel(caminho_xlsx, header=None)
    df = df.iloc[6:].reset_index(drop=True)
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all").reset_index(drop=True)
    return df


def obter_acumuladores_da_empresa(df_empresa: pd.DataFrame) -> set[int]:
    acumuladores: set[int] = set()
    if df_empresa.empty:
        return acumuladores

    primeira_coluna = df_empresa.iloc[:, 0].tolist()
    for valor in primeira_coluna:
        codigo = normalizar_codigo(valor)
        if codigo is not None:
            acumuladores.add(codigo)
    return acumuladores


# ============================================================
# NAVEGACAO POR ATALHO
# ============================================================
def entrar_em_acumuladores_por_atalho(agente: AgenteDominio) -> None:
    LOG.info("[INFO] Entrando em Acumuladores via ALT segurado + A + A")
    agente.sequencia_alt_aa(seg_entre_teclas=0.2, seg_final=CFG.tempo_dominio_max)


# ============================================================
# RPA PARA COLETA DOS RELATORIOS
# ============================================================
def coletar_relatorio_empresa(agente: AgenteDominio, codigo_empresa: int) -> None:
    executar_com_retry(
        lambda: agente.pressionar("f8", CFG.tempo_dominio_baixo),
        descricao=f"F8 empresa {codigo_empresa}",
    )
    executar_com_retry(
        lambda: agente.escrever(str(codigo_empresa), CFG.tempo_dominio_baixo),
        descricao=f"Digitacao empresa {codigo_empresa}",
    )
    executar_com_retry(
        lambda: agente.pressionar("enter", CFG.tempo_dominio_baixo),
        descricao=f"Entrada empresa {codigo_empresa}",
    )

    entrar_em_acumuladores_por_atalho(agente)

    agente.clicar_imagem_ou_coordenada(
        "listagem.png",
        CFG.x_listagem,
        CFG.y_listagem,
        CFG.tempo_dominio,
        descricao="Listagem",
    )
    agente.clicar_imagem_ou_coordenada(
        "tirar_relatorio.png",
        CFG.x_tirar_relatorio,
        CFG.y_tirar_relatorio,
        CFG.tempo_dominio_baixo,
        descricao="Tirar relatorio",
    )
    agente.clicar_imagem_ou_coordenada(
        "opcoes_relatorio.png",
        CFG.x_opcoes_relatorio,
        CFG.y_opcoes_relatorio,
        CFG.tempo_dominio_baixo,
        descricao="Opcoes relatorio",
    )
    agente.clicar_imagem_ou_coordenada(
        "trocar_opcao_relatorio.png",
        CFG.x_trocar_opcao,
        CFG.y_trocar_opcao,
        CFG.tempo_dominio_baixo,
        descricao="Trocar opcao",
    )
    agente.clicar_imagem_ou_coordenada(
        "gerar_relatorio.png",
        CFG.x_gerar,
        CFG.y_gerar,
        CFG.tempo_dominio_baixo,
        descricao="Gerar",
    )

    agente.clicar_imagem_ou_coordenada(
        "excel.png",
        CFG.x_excel,
        CFG.y_excel,
        CFG.tempo_dominio_baixo,
        descricao="Excel",
    )
    agente.clicar_imagem_ou_coordenada(
        "nome_arquivo.png",
        CFG.x_nome_arquivo,
        CFG.y_nome_arquivo,
        CFG.tempo_dominio_baixo,
        descricao="Nome arquivo",
    )
    agente.escrever(f" - {codigo_empresa}", CFG.tempo_dominio_baixo)
    
    # Clica em Salvar (Espera 11 segundos - tempo_dominio - para gerar e abrir o Excel)
    agente.clicar_imagem_ou_coordenada(
        "salvar.png",
        CFG.x_salvar,
        CFG.y_salvar,
        CFG.tempo_dominio,
        descricao="Salvar",
    )

    # --- NOVAS ETAPAS INSERIDAS AQUI ---
    LOG.info("[INFO] Fechando Excel aberto automaticamente (Alt + F4)")
    agente.hotkey("alt", "f4", seg=CFG.tempo_dominio_baixo)

    LOG.info("[INFO] Voltando o foco para a aba do Dominio")
    agente.clicar_imagem_ou_coordenada(
        "voltar_dominio.png",
        CFG.x_voltar_dominio,
        CFG.y_voltar_dominio,
        CFG.tempo_dominio_baixo,
        descricao="Voltar aba Dominio",
    )
    # -----------------------------------

    LOG.info(f"[OK] Relatorio solicitado para empresa {codigo_empresa}")


def coletar_relatorios_rpa(empresas_df: pd.DataFrame) -> None:
    if not CFG.coletar_relatorios_rpa:
        return

    agente = criar_agente()
    LOG.info("[INFO] Coleta RPA com atalho ALT+A habilitada. Posicione o sistema antes do inicio.")
    time.sleep(CFG.segundos_iniciais)

    for _, row in empresas_df.iterrows():
        codigo_empresa = int(row["CODIGO_EMPRESA"])
        LOG.info(f"[INFO] Coletando relatorio da empresa {codigo_empresa}")

        try:
            coletar_relatorio_empresa(agente, codigo_empresa)
        except Exception as exc:
            LOG.info(f"[ERRO] Falha na coleta da empresa {codigo_empresa}: {exc}")
            agente.screenshot_debug(f"erro_etapa1_empresa_{codigo_empresa}")
        finally:
            agente.fechar_sistema_para_loop(CFG.esc_loop_qtd, CFG.esc_loop_intervalo)


# ============================================================
# VALIDACAO E SAIDAS
# ============================================================
def gerar_validados() -> pd.DataFrame:
    garantir_pastas()

    empresas_df = carregar_empresas(CFG.planilha_empresas)
    acumuladores_df = carregar_acumuladores_principais(CFG.planilha_acumuladores)

    LOG.info(f"[INFO] Empresas carregadas: {len(empresas_df)}")
    LOG.info(f"[INFO] Acumuladores lidos da planilha principal: {len(acumuladores_df)}")
    LOG.info(f"[INFO] Acumuladores amarelos: {(acumuladores_df['AMARELO_PLANILHA'] == 'SIM').sum()}")

    coletar_relatorios_rpa(empresas_df)
    converter_xls_para_xlsx(CFG.pasta_relatorios)

    resultados =[]

    for _, empresa_row in empresas_df.iterrows():
        codigo_empresa = int(empresa_row["CODIGO_EMPRESA"])
        nome_empresa = empresa_row["NOME_EMPRESA"]
        caminho_relatorio = localizar_relatorio_empresa(codigo_empresa, CFG.pasta_relatorios)

        if caminho_relatorio is None:
            LOG.info(f"[AVISO] Relatorio nao encontrado para empresa {codigo_empresa}")
            for _, acum_row in acumuladores_df.iterrows():
                resultados.append({
                    "EMPRESA": codigo_empresa,
                    "NOME_EMPRESA": nome_empresa,
                    "ACUMULADOR": int(acum_row["ACUMULADOR"]),
                    "NOME_ACUMULADOR": acum_row["NOME_ACUMULADOR"],
                    "AMARELO_PLANILHA": acum_row["AMARELO_PLANILHA"],
                    "DENTRO_FAIXA_212_1000": acum_row["DENTRO_FAIXA_212_1000"],
                    "EXISTE": "NAO",
                    "POSSIVEL": acum_row["AMARELO_PLANILHA"],
                    "INATIVAR": "NAO",
                    "RELATORIO_EMPRESA": "",
                    "OBS": "Relatorio da empresa nao encontrado",
                })
            continue

        try:
            df_empresa = executar_com_retry(
                lambda: limpar_planilha_empresa(caminho_relatorio),
                descricao=f"Leitura do relatorio {caminho_relatorio.name}",
            )
            acumuladores_empresa = obter_acumuladores_da_empresa(df_empresa)

            LOG.info(
                f"[OK] Empresa {codigo_empresa} | relatorio: {caminho_relatorio.name} | "
                f"acumuladores lidos: {len(acumuladores_empresa)}"
            )

            for _, acum_row in acumuladores_df.iterrows():
                codigo_acumulador = int(acum_row["ACUMULADOR"])
                amarelo = acum_row["AMARELO_PLANILHA"]
                existe = "SIM" if codigo_acumulador in acumuladores_empresa else "NAO"
                possivel = "SIM" if amarelo == "SIM" else "NAO"
                inativar = "SIM" if amarelo == "SIM" and existe == "SIM" else "NAO"

                if amarelo == "SIM" and existe == "SIM":
                    obs = "OK para inativacao"
                elif amarelo == "SIM" and existe == "NAO":
                    obs = "Marcado na planilha, mas nao existe no relatorio da empresa"
                else:
                    obs = "Nao marcado em amarelo na planilha principal"

                resultados.append({
                    "EMPRESA": codigo_empresa,
                    "NOME_EMPRESA": nome_empresa,
                    "ACUMULADOR": codigo_acumulador,
                    "NOME_ACUMULADOR": acum_row["NOME_ACUMULADOR"],
                    "AMARELO_PLANILHA": amarelo,
                    "DENTRO_FAIXA_212_1000": acum_row["DENTRO_FAIXA_212_1000"],
                    "EXISTE": existe,
                    "POSSIVEL": possivel,
                    "INATIVAR": inativar,
                    "RELATORIO_EMPRESA": caminho_relatorio.name,
                    "OBS": obs,
                })

        except Exception as exc:
            LOG.info(f"[ERRO] Falha ao processar a empresa {codigo_empresa}: {exc}")
            for _, acum_row in acumuladores_df.iterrows():
                resultados.append({
                    "EMPRESA": codigo_empresa,
                    "NOME_EMPRESA": nome_empresa,
                    "ACUMULADOR": int(acum_row["ACUMULADOR"]),
                    "NOME_ACUMULADOR": acum_row["NOME_ACUMULADOR"],
                    "AMARELO_PLANILHA": acum_row["AMARELO_PLANILHA"],
                    "DENTRO_FAIXA_212_1000": acum_row["DENTRO_FAIXA_212_1000"],
                    "EXISTE": "NAO",
                    "POSSIVEL": acum_row["AMARELO_PLANILHA"],
                    "INATIVAR": "NAO",
                    "RELATORIO_EMPRESA": caminho_relatorio.name,
                    "OBS": f"Erro ao ler relatorio: {exc}",
                })

    resultado_df = pd.DataFrame(resultados)
    if resultado_df.empty:
        raise ValueError("Nenhum dado foi gerado para VALIDADOS.xlsx")

    resultado_df = resultado_df.sort_values(["EMPRESA", "ACUMULADOR"]).reset_index(drop=True)
    resultado_df.to_excel(CFG.arquivo_validados, index=False)
    resultado_df.to_csv(CFG.arquivo_auditoria_csv, index=False, sep=";", encoding="utf-8-sig")
    aplicar_cores_validados(CFG.arquivo_validados)

    resumo = (
        resultado_df.groupby("EMPRESA")["INATIVAR"]
        .apply(lambda s: (s == "SIM").sum())
        .reset_index(name="QTD_INATIVAR")
    )
    resumo.to_csv(CFG.pasta_logs / "resumo_validacao_por_empresa.csv", index=False, sep=";", encoding="utf-8-sig")

    LOG.info(f"[OK] VALIDADOS gerado em: {CFG.arquivo_validados}")
    LOG.info(f"[OK] Auditoria CSV gerada em: {CFG.arquivo_auditoria_csv}")
    return resultado_df


def aplicar_cores_validados(caminho_arquivo: Path) -> None:
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    verde = PatternFill(start_color=CFG.cor_inativar_rgb, end_color=CFG.cor_inativar_rgb, fill_type="solid")
    amarelo = PatternFill(start_color=CFG.cor_possivel_rgb, end_color=CFG.cor_possivel_rgb, fill_type="solid")
    vermelho = PatternFill(start_color=CFG.cor_nao_existe_rgb, end_color=CFG.cor_nao_existe_rgb, fill_type="solid")

    cabecalho = {cell.value: cell.column for cell in ws[1]}
    col_amarelo = cabecalho["AMARELO_PLANILHA"]
    col_existe = cabecalho["EXISTE"]
    col_inativar = cabecalho["INATIVAR"]

    for row in range(2, ws.max_row + 1):
        amarelo_planilha = ws.cell(row=row, column=col_amarelo).value
        existe = ws.cell(row=row, column=col_existe).value
        inativar = ws.cell(row=row, column=col_inativar).value

        if inativar == "SIM":
            fill = verde
        elif amarelo_planilha == "SIM" and existe == "NAO":
            fill = vermelho
        elif amarelo_planilha == "SIM":
            fill = amarelo
        else:
            fill = None

        if fill is not None:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(caminho_arquivo)


if __name__ == "__main__":
    gerar_validados()